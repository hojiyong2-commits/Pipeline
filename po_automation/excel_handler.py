"""
excel_handler.py — Excel 파일 처리 및 COM 자동화 모듈.

CustomerOrderLines 컬럼을 Tracker로 매핑하고,
win32com을 통해 Tracker의 B7 셀 값을 읽으며 CORB Link를 업데이트합니다.
"""

from __future__ import annotations

import json
import logging
import os
import tempfile
from pathlib import Path
from typing import Optional

import openpyxl  # type: ignore[import]
from openpyxl import load_workbook  # type: ignore[import]

logger = logging.getLogger(__name__)

# Tracker Sheet1의 헤더가 위치한 행 번호 (1-based)
_TRACKER_HEADER_ROW: int = 4

# config에서 필수로 존재해야 하는 키 목록
_REQUIRED_CONFIG_KEYS: tuple[str, ...] = (
    "temp_folder",
    "col_file",
    "tracker_file",
    "socal_path",
    "sg_path",
)

# Excel COM 오류값 문자열 집합
_EXCEL_ERROR_VALUES: frozenset[str] = frozenset(
    {"#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NUM!", "#NULL!"}
)


def _safe_resolve(user_path: str, allowed_root: Path) -> Path:
    """경로 순회 공격 방지: resolve() 후 allowed_root 검증.

    Args:
        user_path: 검증할 상대 또는 절대 경로 문자열.
        allowed_root: 허용된 루트 디렉토리 Path.

    Returns:
        검증된 절대 경로 Path 객체.

    Raises:
        ValueError: 경로가 allowed_root 외부로 탈출하면 raise.
    """
    resolved = (allowed_root / user_path).resolve()
    try:
        resolved.relative_to(allowed_root.resolve())
    except ValueError:
        raise ValueError(
            f"Path traversal detected: '{user_path}' escapes allowed root '{allowed_root}'"
        )
    return resolved


def read_text_with_fallback(path: Path) -> str:
    """utf-8 → cp949 → latin-1 순서 인코딩 fallback으로 텍스트 파일 읽기.

    Args:
        path: 읽을 파일의 Path 객체.

    Returns:
        파일 내용 문자열.

    Raises:
        UnicodeDecodeError: 모든 인코딩으로 읽기 실패 시.
        FileNotFoundError: 파일이 존재하지 않을 때.
    """
    if path is None:
        raise TypeError("path must not be None")
    if not isinstance(path, Path):
        raise TypeError(f"path must be pathlib.Path, got {type(path).__name__}")

    for enc in ("utf-8", "cp949", "latin-1"):
        try:
            return path.read_text(encoding=enc)
        except (UnicodeDecodeError, LookupError):
            continue

    raise UnicodeDecodeError(
        "utf-8", b"", 0, 1, f"Cannot decode {path} with any supported encoding"
    )


def load_config(config_path: str) -> dict:
    """config.json을 읽어 설정 딕셔너리를 반환합니다.

    utf-8 → cp949 → latin-1 fallback으로 파일을 읽습니다.
    5개 필수 키가 모두 존재하는지 검증합니다.

    Args:
        config_path: config.json 파일 경로 문자열.

    Returns:
        파싱된 설정 딕셔너리.

    Raises:
        TypeError: config_path가 None이거나 str이 아닐 때.
        ValueError: config_path가 빈 문자열이거나 1~255자 범위를 벗어날 때.
        FileNotFoundError: 파일이 존재하지 않을 때.
        ValueError: 필수 키가 누락된 경우.
        json.JSONDecodeError: JSON 파싱 실패 시.
    """
    # AL type_valid: None → isinstance 순서 고정
    if config_path is None:
        raise TypeError("config_path must not be None")
    if not isinstance(config_path, str):
        raise TypeError(f"config_path must be str, got {type(config_path).__name__}")
    if len(config_path) == 0:
        raise ValueError("config_path must not be empty string")
    if len(config_path) > 255:
        raise ValueError(f"config_path must be 1~255 chars, got {len(config_path)}")

    path = Path(config_path)
    if not path.exists():
        raise FileNotFoundError(f"config 파일 없음: {config_path}")

    raw_text = read_text_with_fallback(path)

    try:
        config: dict = json.loads(raw_text)
    except json.JSONDecodeError as exc:
        raise json.JSONDecodeError(
            f"config.json 파싱 실패: {exc.msg}", exc.doc, exc.pos
        ) from exc

    # 필수 키 검증
    missing = [k for k in _REQUIRED_CONFIG_KEYS if k not in config]
    if missing:
        raise ValueError(f"config에 필수 키 누락: {missing}")

    logger.info("config 로드 완료: %s", config_path)
    return config


def map_columns(config: dict) -> list[dict]:
    """CustomerOrderLines 컬럼을 Tracker Sheet1에 매핑하여 추가합니다.

    - col_file(Customer Order Lines)을 read_only 모드로 읽습니다.
    - tracker_file Sheet1의 4번째 행을 헤더로 사용합니다.
    - 헤더명 완전 일치(strip 적용) 매핑으로 Tracker에 데이터를 추가합니다.
    - Tracker 마지막 실제 데이터 행 +1부터 값을 채웁니다.
    - tempfile → os.replace() 원자적 패턴으로 저장합니다.

    Args:
        config: 설정 딕셔너리 (col_file, tracker_file 필수).

    Returns:
        추가된 행의 {project_id, location, po, row_index} 목록.

    Raises:
        TypeError: config가 None이거나 dict가 아닐 때.
        ValueError: config가 빈 딕셔너리이거나 필수 키 누락 시.
        FileNotFoundError: col_file 또는 tracker_file이 존재하지 않을 때.
    """
    # AL type_valid: None → isinstance 순서 고정
    if config is None:
        raise TypeError("config must not be None")
    if not isinstance(config, dict):
        raise TypeError(f"config must be dict, got {type(config).__name__}")
    if len(config) == 0:
        raise ValueError("config must not be empty")

    col_path = Path(config["col_file"])
    tracker_path = Path(config["tracker_file"])

    if not col_path.exists():
        raise FileNotFoundError(f"col_file 없음: {col_path}")
    if not tracker_path.exists():
        raise FileNotFoundError(f"tracker_file 없음: {tracker_path}")

    # CustomerOrderLines 읽기 (read_only 모드)
    logger.info("CustomerOrderLines 읽기: %s", col_path)
    col_wb = load_workbook(str(col_path), read_only=True, data_only=True)
    col_ws = col_wb.active

    # CustomerOrderLines 헤더 (1번째 행)
    col_headers: list[str] = []
    col_rows: list[list] = []

    for i, row in enumerate(col_ws.iter_rows(values_only=True)):
        if i == 0:
            col_headers = [str(c).strip() if c is not None else "" for c in row]
        else:
            col_rows.append(list(row))

    col_wb.close()
    logger.info("CustomerOrderLines 행 수: %d", len(col_rows))

    # Tracker 읽기/쓰기 모드
    logger.info("Tracker 파일 읽기: %s", tracker_path)
    tracker_wb = load_workbook(str(tracker_path))
    tracker_ws = tracker_wb["Sheet1"]

    # 4번째 행 헤더 추출 (1-based 인덱스: 행 4)
    tracker_headers: list[str] = []
    for cell in tracker_ws[_TRACKER_HEADER_ROW]:
        val = cell.value
        tracker_headers.append(str(val).strip() if val is not None else "")

    logger.info("Tracker 헤더: %s", tracker_headers)

    # 헤더 → 컬럼 인덱스 매핑 (Tracker 기준)
    tracker_header_index: dict[str, int] = {
        h: i for i, h in enumerate(tracker_headers) if h
    }

    # CustomerOrderLines 헤더 → Tracker 헤더 매핑
    col_to_tracker: dict[int, int] = {}
    for col_idx, col_header in enumerate(col_headers):
        if col_header in tracker_header_index:
            col_to_tracker[col_idx] = tracker_header_index[col_header]

    # Tracker 마지막 실제 데이터 행 탐색 (헤더 행 제외)
    last_data_row = _TRACKER_HEADER_ROW  # 헤더 행
    for row in tracker_ws.iter_rows(
        min_row=_TRACKER_HEADER_ROW + 1, values_only=True
    ):
        if any(cell is not None for cell in row):
            last_data_row += 1

    write_start_row = last_data_row + 1
    logger.info("Tracker 쓰기 시작 행: %d", write_start_row)

    added_rows: list[dict] = []

    for row_offset, col_row in enumerate(col_rows):
        tracker_row_index = write_start_row + row_offset
        row_data: dict[int, object] = {}

        for col_idx, tracker_idx in col_to_tracker.items():
            if col_idx < len(col_row):
                row_data[tracker_idx] = col_row[col_idx]

        # Tracker에 값 기록 (1-based 컬럼 인덱스)
        for tracker_col_idx, value in row_data.items():
            cell = tracker_ws.cell(
                row=tracker_row_index, column=tracker_col_idx + 1, value=value
            )
            _ = cell  # openpyxl cell 참조

        # project_id, location, po 추출 (헤더명 기준)
        def _get_col_val(header_name: str) -> Optional[str]:
            """CustomerOrderLines에서 특정 헤더 컬럼 값 추출."""
            if header_name in col_headers:
                idx = col_headers.index(header_name)
                val = col_row[idx] if idx < len(col_row) else None
                return str(val).strip() if val is not None else ""
            return ""

        added_rows.append(
            {
                "project_id": _get_col_val("Project ID"),
                "location": _get_col_val("Location"),
                "po": _get_col_val("PO"),
                "row_index": tracker_row_index,
            }
        )

    # 원자적 저장: tempfile → os.replace()
    tmp_path = tracker_path.with_suffix(tracker_path.suffix + ".tmp")
    try:
        tracker_wb.save(str(tmp_path))
        os.replace(str(tmp_path), str(tracker_path))
        logger.info("Tracker 저장 완료 (원자적): %s", tracker_path)
    except Exception as exc:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass
        raise RuntimeError(f"Tracker 저장 실패: {tracker_path}") from exc
    finally:
        tracker_wb.close()

    logger.info("추가된 행 수: %d", len(added_rows))
    return added_rows


def read_b7_via_com(tracker_path: str, project_id: str) -> str:
    """win32com을 통해 Tracker Sheet2의 B7 셀 값을 읽습니다.

    Sheet2 A1에 project_id를 입력 후 Application.Calculate()를 호출하고
    B7 값을 문자열로 반환합니다.

    Args:
        tracker_path: Tracker Excel 파일 절대 경로 문자열.
        project_id: Sheet2 A1에 입력할 Project ID 문자열.

    Returns:
        B7 셀 값 문자열.

    Raises:
        TypeError: 인자가 None이거나 str이 아닐 때.
        ValueError: project_id가 빈 문자열이거나 1~255자 범위를 벗어날 때.
        ValueError: B7이 비어있거나 Excel 오류값일 때.
        RuntimeError: COM 통신 오류 시.
    """
    # AL type_valid: None → isinstance 순서 고정
    if tracker_path is None:
        raise TypeError("tracker_path must not be None")
    if not isinstance(tracker_path, str):
        raise TypeError(f"tracker_path must be str, got {type(tracker_path).__name__}")
    if len(tracker_path) == 0:
        raise ValueError("tracker_path must not be empty string")
    if len(tracker_path) > 255:
        raise ValueError(f"tracker_path must be 1~255 chars, got {len(tracker_path)}")

    if project_id is None:
        raise TypeError("project_id must not be None")
    if not isinstance(project_id, str):
        raise TypeError(f"project_id must be str, got {type(project_id).__name__}")
    if len(project_id) == 0:
        raise ValueError("project_id must not be empty string")
    if len(project_id) > 255:
        raise ValueError(f"project_id must be 1~255 chars, got {len(project_id)}")

    try:
        import win32com.client  # type: ignore[import]
    except ImportError as exc:
        raise RuntimeError("win32com 로드 실패 — pywin32 설치 필요") from exc

    xl_app = None
    workbook = None
    try:
        xl_app = win32com.client.Dispatch("Excel.Application")
        xl_app.Visible = False
        xl_app.DisplayAlerts = False

        logger.info("Excel COM 열기: %s", tracker_path)
        workbook = xl_app.Workbooks.Open(str(Path(tracker_path).resolve()))

        sheet2 = workbook.Sheets("Sheet2")

        # A1에 project_id 입력 후 재계산
        sheet2.Range("A1").Value = project_id
        xl_app.Calculate()
        logger.info("Excel 재계산 완료 (project_id=%s)", project_id)

        raw_b7 = sheet2.Range("B7").Value

        # B7 비어있으면 ValueError
        if raw_b7 is None or str(raw_b7).strip() == "":
            raise ValueError(f"B7 값이 비어있음 (project_id={project_id})")

        b7_str = str(raw_b7).strip()

        # Excel 오류값 감지
        if b7_str in _EXCEL_ERROR_VALUES:
            raise ValueError(
                f"B7에 Excel 오류값 포함: '{b7_str}' (project_id={project_id})"
            )

        logger.info("B7 값 읽기 성공: %s", b7_str)
        return b7_str

    except ValueError:
        raise  # 이미 구체적 메시지 포함된 ValueError는 그대로 전파
    except Exception as exc:
        raise RuntimeError(
            f"COM B7 읽기 실패 (project_id='{project_id}', path='{tracker_path}'): {exc}"
        ) from exc
    finally:
        try:
            if workbook is not None:
                workbook.Close(SaveChanges=True)
                logger.info("Excel Workbook 저장 후 닫기 완료")
        except Exception as exc:
            logger.error("Workbook 닫기 실패: %s", exc)
        try:
            if xl_app is not None:
                xl_app.Quit()
                logger.info("Excel Application 종료 완료")
        except Exception as exc:
            logger.error("Excel Application 종료 실패: %s", exc)


def update_corb_link(tracker_path: str, row_index: int, corb_path: str) -> None:
    """Tracker Sheet1의 "CORB Link" 컬럼에 corb_path를 저장합니다.

    4번째 행에서 "CORB Link" 헤더를 탐색하고, 해당 row_index 셀에 값을 기록합니다.
    tempfile → os.replace() 원자적 패턴으로 저장합니다.

    Args:
        tracker_path: Tracker Excel 파일 절대 경로 문자열.
        row_index: 업데이트할 행 번호 (1-based 정수, 헤더 행 이후여야 함).
        corb_path: 저장할 CORB 파일/폴더 경로 문자열.

    Raises:
        TypeError: 인자가 None이거나 타입 불일치 시.
        ValueError: row_index가 0 이하이거나 255자 범위 벗어날 때.
        ValueError: "CORB Link" 헤더를 찾을 수 없을 때.
        FileNotFoundError: tracker_file이 존재하지 않을 때.
        RuntimeError: 파일 저장 실패 시.
    """
    # AL type_valid: None → isinstance 순서 고정
    if tracker_path is None:
        raise TypeError("tracker_path must not be None")
    if not isinstance(tracker_path, str):
        raise TypeError(f"tracker_path must be str, got {type(tracker_path).__name__}")
    if len(tracker_path) == 0:
        raise ValueError("tracker_path must not be empty string")
    if len(tracker_path) > 255:
        raise ValueError(f"tracker_path must be 1~255 chars, got {len(tracker_path)}")

    if row_index is None:
        raise TypeError("row_index must not be None")
    if not isinstance(row_index, int):
        raise TypeError(f"row_index must be int, got {type(row_index).__name__}")
    # row_index는 헤더 행(4) 이후여야 하므로 최소 5 이상
    if row_index <= _TRACKER_HEADER_ROW:
        raise ValueError(
            f"row_index must be > {_TRACKER_HEADER_ROW} (header row), got {row_index}"
        )

    if corb_path is None:
        raise TypeError("corb_path must not be None")
    if not isinstance(corb_path, str):
        raise TypeError(f"corb_path must be str, got {type(corb_path).__name__}")
    if len(corb_path) == 0:
        raise ValueError("corb_path must not be empty string")
    if len(corb_path) > 255:
        raise ValueError(f"corb_path must be 1~255 chars, got {len(corb_path)}")

    path = Path(tracker_path)
    if not path.exists():
        raise FileNotFoundError(f"tracker_file 없음: {tracker_path}")

    wb = load_workbook(str(path))
    ws = wb["Sheet1"]

    # 4번째 행에서 "CORB Link" 헤더 탐색
    corb_col: Optional[int] = None
    for cell in ws[_TRACKER_HEADER_ROW]:
        if cell.value is not None and str(cell.value).strip() == "CORB Link":
            corb_col = cell.column
            break

    if corb_col is None:
        wb.close()
        raise ValueError(
            f"'CORB Link' 헤더를 Sheet1 {_TRACKER_HEADER_ROW}번째 행에서 찾을 수 없음"
        )

    # row_index 셀에 corb_path 기록
    ws.cell(row=row_index, column=corb_col, value=corb_path)
    logger.info("CORB Link 기록: row=%d, col=%d, value=%s", row_index, corb_col, corb_path)

    # 원자적 저장: tempfile → os.replace()
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    try:
        wb.save(str(tmp_path))
        os.replace(str(tmp_path), str(path))
        logger.info("Tracker CORB Link 저장 완료 (원자적): %s", path)
    except Exception as exc:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass
        raise RuntimeError(f"Tracker CORB Link 저장 실패: {path}") from exc
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Self-Verification Block
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import tempfile as _tempfile
    import json as _json

    # --- load_config 정상 케이스 ---
    _cfg = {
        "temp_folder": "C:\\TempPO",
        "col_file": "C:\\col.xlsx",
        "tracker_file": "C:\\tracker.xlsx",
        "socal_path": "C:\\SoCal",
        "sg_path": "C:\\SG",
    }
    with _tempfile.NamedTemporaryFile(
        mode="w", suffix=".json", delete=False, encoding="utf-8"
    ) as f:
        _json.dump(_cfg, f)
        _cfg_path = f.name

    loaded = load_config(_cfg_path)
    assert loaded["temp_folder"] == "C:\\TempPO", "load_config 정상 케이스 실패"
    os.unlink(_cfg_path)

    # --- load_config None 입력 → TypeError ---
    try:
        load_config(None)  # type: ignore[arg-type]
        assert False, "None 입력 예외 미발생"
    except TypeError:
        pass

    # --- load_config 빈 문자열 → ValueError ---
    try:
        load_config("")
        assert False, "빈 문자열 예외 미발생"
    except ValueError:
        pass

    # --- read_b7_via_com 타입 검증 ---
    try:
        read_b7_via_com(None, "PID")  # type: ignore[arg-type]
        assert False, "tracker_path None 예외 미발생"
    except TypeError:
        pass

    try:
        read_b7_via_com("path.xlsx", None)  # type: ignore[arg-type]
        assert False, "project_id None 예외 미발생"
    except TypeError:
        pass

    # --- update_corb_link 타입 검증 ---
    try:
        update_corb_link(None, 5, "path")  # type: ignore[arg-type]
        assert False, "tracker_path None 예외 미발생"
    except TypeError:
        pass

    try:
        update_corb_link("t.xlsx", 4, "path")  # row_index <= header_row → ValueError
        assert False, "row_index 경계값 예외 미발생"
    except ValueError:
        pass

    try:
        update_corb_link("t.xlsx", -1, "path")  # 음수 → ValueError
        assert False, "row_index 음수 예외 미발생"
    except ValueError:
        pass

    print("[SELF-VERIFY] excel_handler.py OK")
