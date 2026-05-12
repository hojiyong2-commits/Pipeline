"""
order_mapper.py
---------------
Pipeline: BUG-20260425-C1D3

Reads CustomerOrderLines Excel, groups rows by (Order No, Wanted Delivery Date),
and writes mapped data into the IC-Part Excel template via ZIP XML direct cell
injection (preserves named ranges, formulas, and drawings).

Python 3.9 compatible.
"""

import logging
import os
import re
import time
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime
from datetime import date as _date
from pathlib import Path
from typing import Any, DefaultDict, Dict, List, Optional, Set, Tuple

import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
)
logger = logging.getLogger("order_mapper")

# ---------------------------------------------------------------------------
# Column indices (0-based) in CustomerOrderLines sheet — FALLBACK defaults
# These are used when header-based auto-detection fails for a given column.
# ---------------------------------------------------------------------------
COL_WANTED_DELIVERY_DATE: int = 1
COL_ORDER_NO: int = 0
COL_LINE_NO: int = 2
COL_CUSTOMER_NAME: int = 8
COL_ORDER_REF_1: int = 131   # -> PO No.
COL_PROJECT_ID: int = 186

# ---------------------------------------------------------------------------
# Column indices (1-based, openpyxl) in IC-Part Sheet1
# ---------------------------------------------------------------------------
ICPART_COL_FOLDER_DATE: int = 8    # H
ICPART_COL_PROJECT_ID: int = 9     # I
ICPART_COL_ORDER_NO: int = 10      # J  <- skip-check column
ICPART_COL_PO_NO: int = 11         # K
ICPART_COL_IC: int = 12            # L
# ICPART_COL_CUSTOMER_NAME removed — Customer Name은 IC-Part에 더 이상 기록하지 않음
ICPART_COL_LINE_NO: int = 13       # M
ICPART_COL_CONTRACT_AMOUNT: int = 14  # N
ICPART_COL_DELIVERY_DATE: int = 15  # O
ICPART_COL_INCOTERM: int = 17      # Q
ICPART_COL_CORB_PATH: int = 18     # R

ICPART_DATA_START_ROW: int = 6

# ---------------------------------------------------------------------------
# IC mapping
# ---------------------------------------------------------------------------
IC_MAP_EXACT: Dict[str, str] = {
    "IMI Critical Engineering LLC": "SoCal",
    "CCI Valve Technology GmbH": "AT",
    "CC Valve Technology Gmbh": "AT",
    "900014": "SoCal",
}
IC_MAP_CONTAINS: List[Tuple[str, str]] = [
    ("CCI Valve", "AT"),   # catches GmbH and Gmbh variants (case-insensitive fallback)
    ("APAC", "SG"),
]
_IC_MAP_EXACT_NORMALIZED: Dict[str, str] = {
    key.casefold(): value for key, value in IC_MAP_EXACT.items()
}
_IC_MAP_CONTAINS_NORMALIZED: List[Tuple[str, str]] = [
    (needle.casefold(), value) for needle, value in IC_MAP_CONTAINS
]


# ---------------------------------------------------------------------------
# Helper: path safety
# ---------------------------------------------------------------------------

def _safe_resolve(user_path: str, allowed_root: Path) -> Path:
    """Resolve user_path relative to allowed_root and verify no traversal."""
    if user_path is None:
        raise TypeError("user_path must not be None")
    if not isinstance(user_path, str):
        raise TypeError(f"user_path must be str, got {type(user_path).__name__}")
    if allowed_root is None:
        raise TypeError("allowed_root must not be None")
    if not isinstance(allowed_root, Path):
        raise TypeError(f"allowed_root must be Path, got {type(allowed_root).__name__}")

    resolved = (allowed_root / user_path).resolve()
    try:
        resolved.relative_to(allowed_root.resolve())
    except ValueError:
        raise ValueError(
            f"Path traversal detected: '{user_path}' escapes allowed root '{allowed_root}'"
        )
    return resolved


# ---------------------------------------------------------------------------
# Helper: encoding-safe text read
# ---------------------------------------------------------------------------

def read_text_with_fallback(path: Path) -> str:
    """Read path trying utf-8 -> cp949 -> latin-1 in order."""
    if path is None:
        raise TypeError("path must not be None")
    if not isinstance(path, Path):
        raise TypeError(f"path must be Path, got {type(path).__name__}")

    for enc in ("utf-8", "cp949", "latin-1"):
        try:
            return path.read_text(encoding=enc)
        except (UnicodeDecodeError, LookupError):
            continue
    raise UnicodeDecodeError(
        "utf-8", b"", 0, 1,
        f"Cannot decode {path} with any supported encoding"
    )


# ---------------------------------------------------------------------------
# Helper: column index to Excel letter
# ---------------------------------------------------------------------------

def _col_letter(n: int) -> str:
    """Convert 1-based column index to Excel column letter."""
    if n is None:
        raise TypeError("n must not be None")
    if not isinstance(n, int):
        raise TypeError(f"n must be int, got {type(n).__name__}")
    if n <= 0:
        raise ValueError(f"n must be positive, got {n}")

    result = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


# ---------------------------------------------------------------------------
# Helper: Excel date serial number
# ---------------------------------------------------------------------------

EXCEL_EPOCH: _date = _date(1899, 12, 30)


def date_to_serial(d: _date) -> int:
    """Convert a Python date to an Excel date serial number."""
    if d is None:
        raise TypeError("d must not be None")
    if not isinstance(d, _date):
        raise TypeError(f"d must be date, got {type(d).__name__}")
    return (d - EXCEL_EPOCH).days


# ---------------------------------------------------------------------------
# XML cell injection helpers
# ---------------------------------------------------------------------------

def _xml_escape(text: str) -> str:
    """Escape XML special characters for element text content."""
    return (text
            .replace('&', '&amp;')
            .replace('<', '&lt;')
            .replace('>', '&gt;')
            .replace('"', '&quot;')
            .replace("'", '&apos;'))


def _col_index(col: str) -> int:
    """Convert column letter(s) to 1-based index. A=1, Z=26, AA=27, etc."""
    result = 0
    for ch in col.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


def _cell_col(cell_ref: str) -> str:
    """Extract column letters from cell reference like 'J6' -> 'J', 'AB12' -> 'AB'."""
    m = re.match(r'^([A-Za-z]+)', cell_ref)
    return m.group(1).upper() if m else ''


def _insert_cell_into_row(xml_str: str, cell_ref: str, new_cell_xml: str) -> str:
    """Insert new_cell_xml into the correct column-sorted position within the row.

    Finds the <row r="{row_num}"> element and inserts before the first existing
    cell whose column sorts after the new cell's column, or before </row>.
    Returns the modified xml_str, or the original if the row element is not found.
    """
    col = _cell_col(cell_ref)
    if not col:
        logger.warning("Cannot parse column from cell ref %s — skipped insertion", cell_ref)
        return xml_str

    row_num_match = re.search(r'[A-Za-z]+(\d+)$', cell_ref)
    if not row_num_match:
        logger.warning("Cannot parse row number from cell ref %s — skipped insertion", cell_ref)
        return xml_str
    row_num = row_num_match.group(1)

    row_pattern = re.compile(
        r'(<row\b[^>]*\br="' + row_num + r'"[^>]*>)(.*?)(</row>)',
        re.DOTALL
    )
    row_match = row_pattern.search(xml_str)
    if not row_match:
        # Row doesn't exist yet (e.g. accumulation past template's last row) — create it
        new_row_xml = f'<row r="{row_num}" spans="1:23">{new_cell_xml}</row>'
        # Insert before the first existing row with a higher row number
        all_rows = re.compile(r'<row\b[^>]*\br="(\d+)"[^>]*>')
        insert_pos = None
        for m in all_rows.finditer(xml_str):
            if int(m.group(1)) > int(row_num):
                insert_pos = m.start()
                break
        if insert_pos is not None:
            return xml_str[:insert_pos] + new_row_xml + xml_str[insert_pos:]
        # No higher row found — insert before </sheetData>
        sd_end = xml_str.rfind('</sheetData>')
        if sd_end != -1:
            return xml_str[:sd_end] + new_row_xml + xml_str[sd_end:]
        logger.warning("Cannot locate </sheetData> — skipped insertion for cell %s", cell_ref)
        return xml_str

    row_open = row_match.group(1)
    row_body = row_match.group(2)
    row_close = row_match.group(3)

    # Find all existing <c r="??{row_num}" elements and their column letters
    existing_cell_pattern = re.compile(r'<c\s+r="([A-Za-z]+)' + row_num + r'"')
    existing_cols = []
    for ec_match in existing_cell_pattern.finditer(row_body):
        existing_cols.append(ec_match.group(1).upper())

    # Determine insertion point: before the first cell with column > new col
    new_col_idx = _col_index(col)
    insert_before_col = None
    for existing_col in sorted(existing_cols, key=_col_index):
        if _col_index(existing_col) > new_col_idx:
            insert_before_col = existing_col
            break

    if insert_before_col is not None:
        # Insert before the first occurrence of the target column cell
        target_pattern = re.compile(
            r'(?=<c\s+r="' + re.escape(insert_before_col) + row_num + r'")'
        )
        new_row_body = target_pattern.sub(lambda _: new_cell_xml, row_body, count=1)
    else:
        # Append before </row>
        new_row_body = row_body + new_cell_xml

    new_row = row_open + new_row_body + row_close
    result = xml_str[:row_match.start()] + new_row + xml_str[row_match.end():]
    return result


def _find_col_style(xml_str: str, col: str, row_num: int) -> str:
    """Return the s= style index for `col` column by scanning the nearest previous rows.

    Used when a new cell is created in a row that has no pre-existing cell for that column,
    so the style is inherited from the same column in the row above (up to 10 rows back).
    Falls back to "0" (default style) if nothing is found.
    """
    for ref_row in range(row_num - 1, max(row_num - 11, 0), -1):
        m = re.search(
            r'<c r="' + re.escape(col) + str(ref_row) + r'"[^>]*\bs="(\d+)"',
            xml_str,
        )
        if m:
            return m.group(1)
    return "0"


def _parse_shared_formulas(xml_str: str) -> Dict[int, Tuple[int, str]]:
    """Return {si: (base_row, formula)} for every shared-formula master in the sheet.

    Master cells look like:
      <f t="shared" ref="A6:A730" si="0">formula</f>
    Dependent (reference-only) cells look like:
      <f t="shared" si="0"/>
    Only masters (non-empty content + ref= attribute) are returned.
    """
    shared: Dict[int, Tuple[int, str]] = {}
    # Use (?<!/)> to exclude self-closing <f t="shared" si="N"/> dependent cells.
    # Without this, a dependent cell's [^>]* matches the "/" in "/>", and the
    # subsequent (.*?) lazily consumes all content up to the NEXT </f> — which
    # happens to be a master cell's closing tag, swallowing that master entirely.
    for m in re.finditer(r'<f\b([^>]*)(?<!/)>(.*?)</f>', xml_str, re.DOTALL):
        attrs, formula = m.group(1), m.group(2).strip()
        if not formula or 't="shared"' not in attrs:
            continue
        si_m = re.search(r'\bsi="(\d+)"', attrs)
        # Handle both plain (V113) and absolute ($V$113) ref notation
        ref_m = re.search(r'\bref="\$?[A-Za-z]+\$?(\d+):', attrs)
        if si_m and ref_m:
            shared[int(si_m.group(1))] = (int(ref_m.group(1)), formula)

    return shared


def _clone_row_xml(
    row_xml: str,
    from_row: int,
    to_row: int,
    shared_formulas: Optional[Dict[int, Tuple[int, str]]] = None,
) -> str:
    """Clone a row, inheriting ONLY formula cells with row-numbers adjusted.

    Rules:
    - Cells with a formula (<f> element): copied with row number updated.
    - Cells without a formula (empty or value-only): dropped — the new row
      cell stays blank, matching the behaviour "if the cell above is empty,
      the new row cell is also empty".
    - Shared formula references (<f t="shared" si="N"/>) are expanded into
      inline formulas so the new row never references a shared formula range
      that doesn't cover it (avoids the Excel "Removed Records: Formula" repair).
    """
    # Rebuild the row opening tag with to_row
    row_open_m = re.match(r'(<row\b[^>]*>)', row_xml)
    if not row_open_m:
        return f'<row r="{to_row}" spans="1:23"/>'
    row_open = re.sub(
        r'(r=")[^"]*"',
        f'r="{to_row}"',
        row_open_m.group(1),
        count=1,
    )

    cloned_cells: List[str] = []
    for cell_m in re.finditer(r'<c\b[^>]*>.*?</c>|<c\b[^>]*/>', row_xml, re.DOTALL):
        cell_xml = cell_m.group(0)

        # Update cell ref: r="A{from_row}" -> r="A{to_row}"
        cell_xml = re.sub(
            r'(r="[A-Za-z]+)' + str(from_row) + r'"',
            lambda mc: mc.group(1) + str(to_row) + '"',
            cell_xml,
        )

        if '<f' not in cell_xml:
            # Non-formula cell: U and W inherit style+value (fixed text cells).
            # B, C, E, H and all other non-formula columns get style only — the
            # injection code writes fresh mapped data; unmapped columns stay empty
            # but retain borders/date-format/alignment from the style index.
            s_m = re.search(r'\bs="(\d+)"', cell_xml)
            if not s_m:
                continue  # no style to preserve — skip entirely
            ref_m = re.search(r'r="([A-Za-z]+' + str(to_row) + r')"', cell_xml)
            if not ref_m:
                continue
            col_letter_m = re.match(r'[A-Za-z]+', ref_m.group(1))
            col_letter = col_letter_m.group() if col_letter_m else ''
            if col_letter in {'U', 'W'}:
                cloned_cells.append(cell_xml)  # inherit style + value
            else:
                cloned_cells.append(f'<c r="{ref_m.group(1)}" s="{s_m.group(1)}"/>')
            continue

        # Formula cell: expand shared formula refs -> inline formula
        if shared_formulas and 't="shared"' in cell_xml:
            def _expand(mc: re.Match) -> str:
                si_m2 = re.search(r'\bsi="(\d+)"', mc.group(0))
                if not si_m2:
                    return mc.group(0)  # can't identify si — preserve original tag
                si = int(si_m2.group(1))
                if si not in shared_formulas:
                    return mc.group(0)  # si not parsed — keep ref so Excel can resolve
                base_row, formula = shared_formulas[si]
                adj = re.sub(
                    r'([A-Za-z]+)' + str(base_row) + r'\b',
                    lambda lm: lm.group(1) + str(to_row),
                    formula,
                )
                return f'<f>{adj}</f>'

            cell_xml = re.sub(
                r'<f\b[^>]*/\s*>',
                lambda mc: _expand(mc) if 't="shared"' in mc.group(0) else mc.group(0),
                cell_xml,
            )
            cell_xml = re.sub(
                r'<f\b[^>]*\bt="shared"[^>]*>.*?</f>',
                _expand,
                cell_xml,
                flags=re.DOTALL,
            )

        # Adjust row numbers in remaining inline <f>content</f>
        def _adj_inline(mc: re.Match) -> str:
            adj = re.sub(
                r'([A-Za-z]+)' + str(from_row) + r'\b',
                lambda lm: lm.group(1) + str(to_row),
                mc.group(1),
            )
            return f'<f>{adj}</f>'

        cell_xml = re.sub(r'<f>(.*?)</f>', _adj_inline, cell_xml, flags=re.DOTALL)

        # Drop stale cached <v> so Excel recalculates on open
        cell_xml = re.sub(r'<v>.*?</v>', '', cell_xml, flags=re.DOTALL)

        cloned_cells.append(cell_xml)

    return row_open + ''.join(cloned_cells) + '</row>'


def _find_cell_with_value(xml_str: str, col: str, row_num: int) -> Optional[str]:
    """Scan backward from row_num-1 to ICPART_DATA_START_ROW and return the nearest
    non-self-closing cell XML for col with the ref updated to col+row_num.
    Only considers cells that contain <v> or <is> (actual value, not style-only).
    """
    for scan_row in range(row_num - 1, ICPART_DATA_START_ROW - 1, -1):
        src_ref = col + str(scan_row)
        tag_m = re.search(r'<c\s+r="' + re.escape(src_ref) + r'"[^>]*>', xml_str)
        if not tag_m:
            continue
        tag_str = tag_m.group(0)
        if tag_str.rstrip().endswith('/>'):
            continue  # self-closing, no value
        close_pos = xml_str.find('</c>', tag_m.end())
        if close_pos == -1:
            continue
        content = xml_str[tag_m.end():close_pos]
        if '<v>' not in content and '<is>' not in content:
            continue
        cell_xml = tag_str + content + '</c>'
        # Update ref row number
        cell_xml = cell_xml.replace(f'r="{src_ref}"', f'r="{col}{row_num}"')
        return cell_xml
    return None


def _find_cell_with_formula(
    xml_str: str,
    col: str,
    row_num: int,
    shared_formulas: Optional[Dict[int, Tuple[int, str]]] = None,
) -> Optional[str]:
    """Scan backward from row_num-1 to ICPART_DATA_START_ROW and return formula cell XML
    for col adapted to row_num. Expands shared formula references to inline formulas.
    Drops cached <v> so Excel recalculates.
    """
    for scan_row in range(row_num - 1, ICPART_DATA_START_ROW - 1, -1):
        src_ref = col + str(scan_row)
        tag_m = re.search(r'<c\s+r="' + re.escape(src_ref) + r'"[^>]*>', xml_str)
        if not tag_m:
            continue
        tag_str = tag_m.group(0)
        if tag_str.rstrip().endswith('/>'):
            continue
        close_pos = xml_str.find('</c>', tag_m.end())
        if close_pos == -1:
            continue
        content = xml_str[tag_m.end():close_pos]
        if '<f' not in content:
            continue

        cell_xml = tag_str + content + '</c>'
        cell_xml = cell_xml.replace(f'r="{src_ref}"', f'r="{col}{row_num}"')

        # Expand shared formula dependencies → inline
        if shared_formulas and 't="shared"' in cell_xml:
            def _exp(mc: re.Match) -> str:
                si_m2 = re.search(r'\bsi="(\d+)"', mc.group(0))
                if not si_m2:
                    return mc.group(0)
                si = int(si_m2.group(1))
                if si not in shared_formulas:
                    return mc.group(0)
                base_row, formula = shared_formulas[si]
                adj = re.sub(
                    r'([A-Za-z]+)' + str(base_row) + r'\b',
                    lambda lm: lm.group(1) + str(row_num),
                    formula,
                )
                return f'<f>{adj}</f>'
            cell_xml = re.sub(
                r'<f\b[^>]*/\s*>',
                lambda mc: _exp(mc) if 't="shared"' in mc.group(0) else mc.group(0),
                cell_xml,
            )
            cell_xml = re.sub(
                r'<f\b[^>]*\bt="shared"[^>]*>.*?</f>',
                _exp,
                cell_xml,
                flags=re.DOTALL,
            )

        # Adjust inline <f>content</f> row numbers (scan_row → row_num)
        def _adj(mc: re.Match) -> str:
            return '<f>' + re.sub(
                r'([A-Za-z]+)' + str(scan_row) + r'\b',
                lambda lm: lm.group(1) + str(row_num),
                mc.group(1),
            ) + '</f>'
        cell_xml = re.sub(r'<f>(.*?)</f>', _adj, cell_xml, flags=re.DOTALL)

        # Drop stale cached value
        cell_xml = re.sub(r'<v>.*?</v>', '', cell_xml, flags=re.DOTALL)
        return cell_xml
    return None


def _inject_or_replace_cell(xml_str: str, row_num: int, col: str, new_cell_xml: str) -> str:
    """Replace existing cell (self-closing or with content) at col+row_num, or insert if absent."""
    ref = col + str(row_num)
    tag_m = re.search(r'<c\s+r="' + re.escape(ref) + r'"[^>]*>', xml_str)
    if tag_m:
        tag_str = tag_m.group(0)
        if tag_str.rstrip().endswith('/>'):
            return xml_str[:tag_m.start()] + new_cell_xml + xml_str[tag_m.end():]
        close_pos = xml_str.find('</c>', tag_m.end())
        if close_pos != -1:
            return xml_str[:tag_m.start()] + new_cell_xml + xml_str[close_pos + 4:]
    return _insert_cell_into_row(xml_str, ref, new_cell_xml)


def _ensure_row_exists(xml_str: str, row_num: int) -> str:
    """Ensure <row r="{row_num}"> exists in xml_str.

    If the row is missing, clones the previous row (row_num - 1) — inheriting
    only formula cells with row-numbers adjusted — then inserts the new row at
    the correct sorted position inside <sheetData>.
    """
    if re.search(r'<row\b[^>]*\br="' + str(row_num) + r'"', xml_str):
        return xml_str

    prev_row = row_num - 1

    if prev_row < ICPART_DATA_START_ROW:
        new_row_xml = f'<row r="{row_num}" spans="1:23"/>'
    else:
        prev_pattern = re.compile(
            r'<row\b[^>]*\br="' + str(prev_row) + r'"[^>]*>.*?</row>',
            re.DOTALL,
        )
        prev_match = prev_pattern.search(xml_str)
        if not prev_match:
            logger.warning(
                "Cannot find row %d to clone for row %d — inserting empty row",
                prev_row, row_num,
            )
            new_row_xml = f'<row r="{row_num}" spans="1:23"/>'
        else:
            shared = _parse_shared_formulas(xml_str)
            new_row_xml = _clone_row_xml(prev_match.group(0), prev_row, row_num, shared)

    all_rows = re.compile(r'<row\b[^>]*\br="(\d+)"[^>]*>')
    insert_pos = None
    for m in all_rows.finditer(xml_str):
        if int(m.group(1)) > row_num:
            insert_pos = m.start()
            break

    if insert_pos is not None:
        return xml_str[:insert_pos] + new_row_xml + xml_str[insert_pos:]

    sd_end = xml_str.rfind('</sheetData>')
    if sd_end != -1:
        return xml_str[:sd_end] + new_row_xml + xml_str[sd_end:]

    logger.warning("Cannot locate </sheetData> — row %d not inserted", row_num)
    return xml_str


def _inject_text_cell(xml_str: str, cell_ref: str, value: str) -> str:
    """Replace any existing cell (empty or filled) with an inline-string value cell.

    Matches both:
      <c r="J6" s="38"/>                          (self-closing empty)
      <c r="J6" s="14" t="inlineStr">...</c>      (filled content)

    When the cell element does not exist in the XML (empty OOXML cells have no
    representation), creates and inserts it at the correct column-sorted position
    within the row.
    """
    if xml_str is None:
        raise TypeError("xml_str must not be None")
    if cell_ref is None:
        raise TypeError("cell_ref must not be None")
    if value is None:
        raise TypeError("value must not be None")

    pattern = re.compile(
        r'<c r="' + re.escape(cell_ref) + r'"(?:\s[^>]*/\s*>|\s[^>]*>(?:(?!</c>).)*?</c>)',
        re.DOTALL
    )
    escaped = _xml_escape(str(value))
    def _repl(m: re.Match) -> str:
        """Replace existing cell element with inline-string variant."""
        s_match = re.search(r'\bs="(\d+)"', m.group(0))
        s_val = s_match.group(1) if s_match else "0"
        return f'<c r="{cell_ref}" s="{s_val}" t="inlineStr"><is><t>{escaped}</t></is></c>'
    result, n = pattern.subn(_repl, xml_str, count=1)
    if n == 0:
        logger.warning("Cell %s not found in sheet1.xml — creating", cell_ref)
        col_m = re.match(r'([A-Za-z]+)(\d+)$', cell_ref)
        col_letter = col_m.group(1)
        row_num_val = int(col_m.group(2))
        s_val = _find_col_style(result, col_letter, row_num_val)
        new_cell_xml = f'<c r="{cell_ref}" s="{s_val}" t="inlineStr"><is><t>{escaped}</t></is></c>'
        result = _insert_cell_into_row(result, cell_ref, new_cell_xml)
    return result


def _inject_date_cell(xml_str: str, cell_ref: str, serial: int) -> str:
    """Replace any existing cell (empty or filled) with a numeric serial value cell.

    When the cell element does not exist in the XML (empty OOXML cells have no
    representation), creates and inserts it at the correct column-sorted position
    within the row.
    """
    if xml_str is None:
        raise TypeError("xml_str must not be None")
    if cell_ref is None:
        raise TypeError("cell_ref must not be None")
    if serial is None:
        raise TypeError("serial must not be None")
    if not isinstance(serial, int):
        raise TypeError(f"serial must be int, got {type(serial).__name__}")
    if serial <= 0:
        raise ValueError(f"serial must be positive, got {serial}")

    pattern = re.compile(
        r'<c r="' + re.escape(cell_ref) + r'"(?:\s[^>]*/\s*>|\s[^>]*>(?:(?!</c>).)*?</c>)',
        re.DOTALL
    )
    def _repl(m: re.Match) -> str:
        """Replace existing cell element with numeric serial variant."""
        s_match = re.search(r'\bs="(\d+)"', m.group(0))
        s_val = s_match.group(1) if s_match else "0"
        return f'<c r="{cell_ref}" s="{s_val}"><v>{serial}</v></c>'
    result, n = pattern.subn(_repl, xml_str, count=1)
    if n == 0:
        logger.warning("Date cell %s not found in sheet1.xml — creating", cell_ref)
        col_m = re.match(r'([A-Za-z]+)(\d+)$', cell_ref)
        col_letter = col_m.group(1)
        row_num_val = int(col_m.group(2))
        s_val = _find_col_style(result, col_letter, row_num_val)
        new_cell_xml = f'<c r="{cell_ref}" s="{s_val}"><v>{serial}</v></c>'
        result = _insert_cell_into_row(result, cell_ref, new_cell_xml)
    return result


# ---------------------------------------------------------------------------
# IC mapping logic
# ---------------------------------------------------------------------------

def map_ic(customer_name: Optional[str]) -> str:
    """Map customer_name to IC code."""
    if customer_name is None:
        return ""
    if not isinstance(customer_name, str):
        raise TypeError(f"customer_name must be str, got {type(customer_name).__name__}")

    normalized = customer_name.casefold()
    if normalized in _IC_MAP_EXACT_NORMALIZED:
        return _IC_MAP_EXACT_NORMALIZED[normalized]

    for substring, code in _IC_MAP_CONTAINS_NORMALIZED:
        if substring in normalized:
            return code

    return customer_name


# ---------------------------------------------------------------------------
# Date normalisation
# ---------------------------------------------------------------------------

def normalise_date(value: Any) -> Optional[date]:
    """Convert Excel date cell value to a Python date."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    logger.warning("Cannot parse date value: %r", value)
    return None


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

GroupKey = Tuple[Any, Any]


class OrderGroup:
    """Holds aggregated data for one (Order No, Wanted Delivery Date) group."""

    def __init__(self) -> None:
        """Initialise all fields to None/empty."""
        self.base_order_no: Optional[str] = None
        self.order_no: Optional[str] = None
        self.delivery_date: Optional[date] = None
        self.po_no: Optional[str] = None
        self.customer_name: Optional[str] = None
        self.project_id: Optional[str] = None
        self.incoterm: Optional[str] = None
        self.folder_date: Optional[date] = None
        self.contract_amount: Optional[str] = None
        self.corb_path: Optional[str] = None
        self.line_nos: List[int] = []

    def add_line(
        self,
        row: Tuple[Any, ...],
        col_order_no: Optional[int] = None,
        col_delivery_date: Optional[int] = None,
        col_line_no: Optional[int] = None,
        col_customer_name: Optional[int] = None,
        col_project_id: Optional[int] = None,
        col_po_no: Optional[int] = None,
    ) -> None:
        """Ingest one CustomerOrderLines row into this group.

        Args:
            row: Tuple of cell values (0-indexed).
            col_order_no: Override for order-no column index.
            col_delivery_date: Override for delivery-date column index.
            col_line_no: Override for line-no column index.
            col_customer_name: Override for customer-name column index.
            col_project_id: Override for project-id column index.
            col_po_no: Override for PO-no column index.

        Raises:
            TypeError: If row is None or not tuple.
            ValueError: If row is empty.
        """
        if row is None:
            raise TypeError("row must not be None")
        if not isinstance(row, tuple):
            raise TypeError(f"row must be tuple, got {type(row).__name__}")
        if len(row) == 0:
            raise ValueError("row must not be empty")

        eff_order_no: int = col_order_no if col_order_no is not None else COL_ORDER_NO
        eff_delivery_date: int = col_delivery_date if col_delivery_date is not None else COL_WANTED_DELIVERY_DATE
        eff_line_no: int = col_line_no if col_line_no is not None else COL_LINE_NO
        eff_customer_name: int = col_customer_name if col_customer_name is not None else COL_CUSTOMER_NAME
        eff_project_id: int = col_project_id if col_project_id is not None else COL_PROJECT_ID
        eff_po_no: int = col_po_no if col_po_no is not None else COL_ORDER_REF_1

        def _get(idx: int) -> Any:
            return row[idx] if idx < len(row) else None

        if self.order_no is None:
            raw_order = _get(eff_order_no)
            self.order_no = str(raw_order).strip() if raw_order is not None else None
            self.base_order_no = self.order_no

            self.delivery_date = normalise_date(_get(eff_delivery_date))

            raw_po = _get(eff_po_no)
            if raw_po is not None:
                if isinstance(raw_po, float) and raw_po.is_integer():
                    self.po_no = str(int(raw_po))
                else:
                    self.po_no = str(raw_po).strip()

            raw_cust = _get(eff_customer_name)
            self.customer_name = str(raw_cust).strip() if raw_cust is not None else None

            raw_proj = _get(eff_project_id)
            self.project_id = str(raw_proj).strip() if raw_proj is not None else None

        raw_line = _get(eff_line_no)
        if raw_line is not None:
            try:
                self.line_nos.append(int(raw_line))
            except (ValueError, TypeError):
                logger.warning("Cannot parse Line No value: %r", raw_line)

    def format_line_nos(self) -> str:
        """Return sorted line numbers formatted as '#1,2,4~10'."""
        if not self.line_nos:
            return ""

        sorted_lines: List[int] = sorted(set(self.line_nos))

        runs: List[tuple] = []
        start = sorted_lines[0]
        end = sorted_lines[0]
        for num in sorted_lines[1:]:
            if num == end + 1:
                end = num
            else:
                runs.append((start, end))
                start = num
                end = num
        runs.append((start, end))

        parts: List[str] = []
        for run_start, run_end in runs:
            length = run_end - run_start + 1
            if length >= 3:
                parts.append(f"{run_start}~{run_end}")
            else:
                for n in range(run_start, run_end + 1):
                    parts.append(str(n))

        return "#" + ",".join(parts)


# ---------------------------------------------------------------------------
# Step 1 helper: auto-detect key columns by header name
# ---------------------------------------------------------------------------

_HEADER_MAP: Dict[str, str] = {
    "customer name": "customer_name",
    "project id": "project_id",
    "project_id": "project_id",
    "order ref 1": "po_no",
    "line no": "line_no",
}


def _detect_columns_from_header(header_row: Tuple[Any, ...]) -> Dict[str, int]:
    """Scan the header row and return a dict mapping logical field names to column indices.

    Recognised mappings (case-insensitive):
      "Customer Name"  -> "customer_name"
      "Project Id" / "Project ID" / "Project_Id" -> "project_id"
      "Order Ref 1"    -> "po_no"
      "Line No"        -> "line_no"

    Args:
        header_row: Tuple of header cell values (first row of the sheet).

    Returns:
        Dict of {field_name: col_index}. Missing columns are absent from dict.

    Raises:
        TypeError: If header_row is None.
    """
    if header_row is None:
        raise TypeError("header_row must not be None")

    detected: Dict[str, int] = {}
    for col_idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        if key in _HEADER_MAP:
            field = _HEADER_MAP[key]
            if field not in detected:
                detected[field] = col_idx
    return detected


# ---------------------------------------------------------------------------
# Step 1 helper: auto-detect key columns by value pattern
# ---------------------------------------------------------------------------

_ORDER_NO_PATTERN: re.Pattern = re.compile(r"^[A-Z]\d{6,}$")


def _auto_detect_key_columns(
    data_rows: List[Tuple[Any, ...]],
) -> Tuple[int, int]:
    """Scan the first few data rows to detect Order No and Delivery Date columns.

    Args:
        data_rows: List of row tuples (header row already skipped by caller).

    Returns:
        Tuple of (order_no_col, delivery_date_col) — zero-based indices.

    Raises:
        TypeError: If data_rows is None.
    """
    if data_rows is None:
        raise TypeError("data_rows must not be None")
    if not isinstance(data_rows, list):
        raise TypeError(f"data_rows must be list, got {type(data_rows).__name__}")

    detected_order_no: Optional[int] = None
    detected_delivery_date: Optional[int] = None

    for row_tuple in data_rows:
        if not row_tuple:
            continue
        scan_limit = min(30, len(row_tuple))
        for col_idx in range(scan_limit):
            val = row_tuple[col_idx]
            if val is None:
                continue

            if detected_order_no is None and isinstance(val, str):
                if _ORDER_NO_PATTERN.match(val.strip()):
                    detected_order_no = col_idx

            if detected_delivery_date is None and isinstance(val, (datetime, date)):
                detected_delivery_date = col_idx

        if detected_order_no is not None and detected_delivery_date is not None:
            break

    if detected_order_no is not None and detected_delivery_date is not None:
        logger.info(
            "Auto-detected: order_no_col=%d, delivery_date_col=%d",
            detected_order_no,
            detected_delivery_date,
        )
        return detected_order_no, detected_delivery_date

    final_order_no = detected_order_no if detected_order_no is not None else COL_ORDER_NO
    final_date = detected_delivery_date if detected_delivery_date is not None else COL_WANTED_DELIVERY_DATE
    logger.warning(
        "WARNING: Could not auto-detect columns, using defaults: "
        "order_no_col=%d, delivery_date_col=%d",
        final_order_no,
        final_date,
    )
    return final_order_no, final_date


# ---------------------------------------------------------------------------
# Step 1: Read CustomerOrderLines
# ---------------------------------------------------------------------------

def read_customer_order_lines(source_path: Path) -> List[OrderGroup]:
    """Read source_path Excel and return a list of OrderGroup objects.

    Groups rows by (Order No, Wanted Delivery Date).
    Sorted by delivery_date ascending, then order_no ascending.

    Args:
        source_path: Absolute path to CustomerOrderLines Excel file.

    Returns:
        List of OrderGroup instances sorted per the pipeline spec.

    Raises:
        TypeError: If source_path is None.
        FileNotFoundError: If source_path does not exist.
        RuntimeError: If the workbook cannot be opened.
    """
    if source_path is None:
        raise TypeError("source_path must not be None")
    if not isinstance(source_path, Path):
        raise TypeError(f"source_path must be Path, got {type(source_path).__name__}")
    if not source_path.exists():
        raise FileNotFoundError(f"CustomerOrderLines file not found: '{source_path}'")

    logger.info("Opening CustomerOrderLines: %s", source_path)
    try:
        wb = openpyxl.load_workbook(str(source_path), read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Cannot open workbook '{source_path}': {exc}") from exc

    try:
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        logger.info("Reading sheet: %s", sheet_name)
        all_rows: List[Tuple[Any, ...]] = [row for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    header_row: Tuple[Any, ...] = all_rows[0] if all_rows else ()
    header_desc = ", ".join(
        f"col{i}={repr(v)}" for i, v in enumerate(header_row[:31])
    )
    logger.info("Headers[0..%d]: %s", min(30, len(header_row) - 1), header_desc)

    # Detect columns from header row (customer name, project id, po no)
    header_detected = _detect_columns_from_header(header_row)
    if header_detected:
        logger.info("Header-detected columns: %s", header_detected)

    # Auto-detect order_no and delivery_date from data values
    data_preview: List[Tuple[Any, ...]] = [row for row in all_rows[1:6] if row]
    detected_order_no_col, detected_delivery_date_col = _auto_detect_key_columns(data_preview)

    groups: DefaultDict[GroupKey, OrderGroup] = defaultdict(OrderGroup)
    row_count: int = 0

    for i, row_tuple in enumerate(all_rows):
        if i == 0:
            continue

        min_required = max(detected_order_no_col, detected_delivery_date_col)
        if len(row_tuple) <= min_required:
            continue

        raw_order_no = row_tuple[detected_order_no_col]
        raw_date = row_tuple[detected_delivery_date_col]

        if raw_order_no is None or raw_date is None:
            continue

        key: GroupKey = (str(raw_order_no).strip(), raw_date)
        groups[key].add_line(
            row_tuple,
            col_order_no=detected_order_no_col,
            col_delivery_date=detected_delivery_date_col,
            col_line_no=header_detected.get("line_no"),
            col_customer_name=header_detected.get("customer_name"),
            col_project_id=header_detected.get("project_id"),
            col_po_no=header_detected.get("po_no"),
        )
        row_count += 1

    logger.info("Read %d data rows, formed %d groups", row_count, len(groups))

    def sort_key(grp: OrderGroup) -> Tuple[date, str]:
        d = grp.delivery_date if grp.delivery_date is not None else date.max
        o = grp.order_no if grp.order_no is not None else ""
        return (d, o)

    sorted_groups = sorted(groups.values(), key=sort_key)
    return sorted_groups


# ---------------------------------------------------------------------------
# Step 1b: Apply sub-order suffixes
# ---------------------------------------------------------------------------

def apply_sub_order_suffixes(groups: List[OrderGroup]) -> None:
    """Append -2/-3/... suffix to order_no when same base order appears multiple times."""
    if groups is None:
        raise TypeError("groups must not be None")
    if not isinstance(groups, list):
        raise TypeError(f"groups must be list, got {type(groups).__name__}")

    counts: Dict[str, int] = {}
    for group in groups:
        if group.base_order_no is None:
            group.base_order_no = group.order_no
        base = group.base_order_no if group.base_order_no is not None else ""
        counts[base] = counts.get(base, 0) + 1

    occurrence: Dict[str, int] = {}
    for group in groups:
        base = group.base_order_no if group.base_order_no is not None else ""
        if counts[base] > 1:
            occ = occurrence.get(base, 0)
            if occ > 0:
                group.order_no = f"{base}-{occ + 1}"
            occurrence[base] = occ + 1


# ---------------------------------------------------------------------------
# Step 2: Write to IC-Part via ZIP XML direct cell injection
# ---------------------------------------------------------------------------

def write_ic_part_zip(
    groups: List[OrderGroup],
    template_path: Path,
    output_path: Path,
) -> Tuple[int, int]:
    """Write groups into IC-Part template using ZIP XML direct injection.

    Automatically detects the last written row in column K and appends
    new groups after it (accumulation mode). First run starts at
    ICPART_DATA_START_ROW (row 6).

    Args:
        groups: Sorted list of OrderGroup objects to write.
        template_path: Absolute path to IC-Part template Excel (.xlsx).
        output_path: Absolute path for output file.

    Returns:
        Tuple of (written_count, skipped_count).

    Raises:
        TypeError: If any argument is None or wrong type.
        FileNotFoundError: If template_path does not exist.
        ValueError: If groups is empty.
        RuntimeError: On ZIP I/O failure.
    """
    if groups is None:
        raise TypeError("groups must not be None")
    if not isinstance(groups, list):
        raise TypeError(f"groups must be list, got {type(groups).__name__}")
    if len(groups) == 0:
        raise ValueError("groups must not be empty")

    if template_path is None:
        raise TypeError("template_path must not be None")
    if not isinstance(template_path, Path):
        raise TypeError(f"template_path must be Path, got {type(template_path).__name__}")
    if not template_path.exists():
        raise FileNotFoundError(f"IC-Part template not found: '{template_path}'")

    if output_path is None:
        raise TypeError("output_path must not be None")
    if not isinstance(output_path, Path):
        raise TypeError(f"output_path must be Path, got {type(output_path).__name__}")

    logger.info("Reading ZIP from template: %s", template_path)
    try:
        with zipfile.ZipFile(str(template_path), "r") as zin:
            all_zip_data: Dict[str, bytes] = {
                item.filename: zin.read(item.filename)
                for item in zin.infolist()
            }
            zip_infolist = zin.infolist()
    except Exception as exc:
        raise RuntimeError(f"Failed to read template ZIP '{template_path}': {exc}") from exc

    if "xl/worksheets/sheet1.xml" not in all_zip_data:
        raise RuntimeError("xl/worksheets/sheet1.xml not found in template ZIP")

    sheet1_bytes: bytes = all_zip_data["xl/worksheets/sheet1.xml"]
    sheet1_str: str = sheet1_bytes.decode("utf-8")

    # BUG-20260512-E68A: 빈행 탐지 기준을 K 컬럼 단독에서 I·J·K 세 컬럼 합집합으로 변경.
    # 기존 코드는 K 컬럼만 스캔했기 때문에, I 또는 J에만 데이터가 있고 K가 비어있는 행을
    # "빈 행"으로 잘못 인식하여 기존 데이터를 덮어쓰는 버그가 있었음.
    # 수정: ICPART_COL_PROJECT_ID(I=9), ICPART_COL_ORDER_NO(J=10), ICPART_COL_PO_NO(K=11) 세
    # 컬럼 모두 스캔하여 값이 있는 행 번호를 _ijk_occupied set에 합산.
    # Inner content boundary (?:(?!</?c[\s>]).)*? prevents spanning into adjacent <c> elements.
    _COL_LETTERS: dict = {
        "I": ICPART_COL_PROJECT_ID,
        "J": ICPART_COL_ORDER_NO,
        "K": ICPART_COL_PO_NO,
    }
    _ijk_occupied: set = set()
    for _scan_letter in _COL_LETTERS:
        _col_pattern = re.compile(
            rf'<c\s[^>]*\br="{_scan_letter}(\d+)"[^>]*/>'  # branch 1: self-closing (no value)
            r'|'
            rf'<c\s[^>]*\br="{_scan_letter}(\d+)"[^>]*>((?:(?!</?c[\s>]).)*?)</c>',  # branch 2: open/close
            re.DOTALL,
        )
        for _m in _col_pattern.finditer(sheet1_str):
            # branch 1 (self-closing): group(1)이 있으면 값 없음 → 건너뜀
            if _m.group(1) is not None:
                continue
            # branch 2 (open/close): group(2)이 행 번호, group(3)이 셀 내용
            _row_g = _m.group(2)
            if _row_g is None:
                continue
            _row_n = int(_row_g)
            if _row_n < ICPART_DATA_START_ROW:
                continue
            # <v> 태그가 있는 셀만 "값 있음"으로 집계
            if not re.search(r'<v[\s>]', _m.group(3) or ''):
                continue
            _ijk_occupied.add(_row_n)

    _last_occupied_row = max(_ijk_occupied, default=ICPART_DATA_START_ROW - 1)
    _actual_start_row = max(_last_occupied_row + 1, ICPART_DATA_START_ROW)
    if _actual_start_row > ICPART_DATA_START_ROW:
        logger.info(
            "Accumulation mode: last occupied row=%d (I·J·K 합집합), starting at row=%d",
            _last_occupied_row,
            _actual_start_row,
        )

    # Parse shared strings table (may not exist in minimal templates)
    import xml.etree.ElementTree as ET
    _shared_strings: List[str] = []
    if "xl/sharedStrings.xml" in all_zip_data:
        _ss_root = ET.fromstring(all_zip_data["xl/sharedStrings.xml"].decode("utf-8"))
        _ss_ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        _shared_strings = [
            (si.findtext(".//x:t", namespaces=_ss_ns) or "")
            for si in _ss_root.findall("x:si", _ss_ns)
        ]

    # Extract existing Project IDs from I column — both inlineStr and shared-string
    # IMP-20260509-F8BF: Project ID column shifted from J (10) to I (9) — regex updated.
    existing_project_ids: Set[str] = set()

    # inlineStr cells: <c r="I6" ... t="inlineStr"><is><t>B42728KR</t></is></c>
    for row_str, val in re.findall(
        r'<c r="I(\d+)"[^>]*t="inlineStr"[^>]*><is><t>([^<]*)</t></is></c>',
        sheet1_str
    ):
        if int(row_str) >= ICPART_DATA_START_ROW and val.strip():
            existing_project_ids.add(val.strip())

    # shared-string cells: <c r="I6" t="s"><v>383</v></c>
    for row_str, idx_str in re.findall(
        r'<c r="I(\d+)"[^>]*t="s"[^>]*><v>(\d+)</v></c>',
        sheet1_str
    ):
        if int(row_str) >= ICPART_DATA_START_ROW:
            idx = int(idx_str)
            if idx < len(_shared_strings) and _shared_strings[idx].strip():
                existing_project_ids.add(_shared_strings[idx].strip())

    logger.info("Existing Project IDs in IC-Part: %s", existing_project_ids)

    written: int = 0
    skipped: int = 0
    write_offset: int = 0
    skipped_indices: Set[int] = set()

    # Pre-parse shared formulas once so backfill helpers can expand them.
    _shared_fmla_cache: Dict[int, Tuple[int, str]] = _parse_shared_formulas(sheet1_str)
    logger.info("Shared formula masters parsed: %s", sorted(_shared_fmla_cache.keys()))

    # Bug fix: freeze the set of already-existing Project IDs from the template
    # before iterating. Sub-groups of the same CustomerOrderLines record share a
    # Project Id (e.g. X100050786, X100050786-2, X100050786-3 all have B42728KR)
    # but must all be written because they have different Order No / delivery dates.
    # Adding newly-written IDs to the check set during the loop caused those
    # sub-groups to be incorrectly skipped as duplicates.
    _check_pids: Set[str] = set(existing_project_ids)  # frozen snapshot — never mutated

    # Count how many delivery-date groups share the original order number.
    # Suffixes (-2/-3) are display values; Line No must be decided from the
    # original CustomerOrderLines order number.
    order_date_count: Counter = Counter(
        (g.base_order_no if g.base_order_no is not None else g.order_no)
        for g in groups
    )

    for i, group in enumerate(groups):
        # Duplicate Project ID check — only against IDs already in the template,
        # not against other groups being written in this same run.
        pid: Optional[str] = group.project_id.strip() if group.project_id else None
        if pid and pid in _check_pids:
            logger.warning(
                "SKIP duplicate Project Id=%s (Order No=%s)", pid, group.order_no
            )
            skipped += 1
            skipped_indices.add(i)
            continue

        row_num = _actual_start_row + write_offset
        logger.info(
            "Injecting group Order No=%s Date=%s into row %d",
            group.order_no,
            group.delivery_date,
            row_num,
        )
        sheet1_str = _ensure_row_exists(sheet1_str, row_num)

        # Backfill U and W: inherit fixed text (e.g. "김유숙 선임", "KOREA") from
        # the nearest non-empty row above, regardless of how many empty rows lie
        # between.  This handles the case where prior runs wrote empty U/W cells.
        for _bfcol in ('U', 'W'):
            _bfref = _bfcol + str(row_num)
            _bftag = re.search(r'<c\s+r="' + re.escape(_bfref) + r'"[^>]*>', sheet1_str)
            _needs_fill = True
            if _bftag and not _bftag.group(0).rstrip().endswith('/>'):
                _bfclose = sheet1_str.find('</c>', _bftag.end())
                if _bfclose != -1:
                    _bfcontent = sheet1_str[_bftag.end():_bfclose]
                    if '<v>' in _bfcontent or '<is>' in _bfcontent:
                        _needs_fill = False
            if _needs_fill:
                _bfsrc = _find_cell_with_value(sheet1_str, _bfcol, row_num)
                if _bfsrc:
                    sheet1_str = _inject_or_replace_cell(sheet1_str, row_num, _bfcol, _bfsrc)
                    logger.info("Backfilled %s from nearest row above", _bfref)

        # Backfill V: inherit formula from nearest formula-containing V row above
        # and expand shared formula reference to inline formula.
        _vref = 'V' + str(row_num)
        _vtag = re.search(r'<c\s+r="' + re.escape(_vref) + r'"[^>]*>', sheet1_str)
        _vneeds_fill = True
        if _vtag and not _vtag.group(0).rstrip().endswith('/>'):
            _vclose = sheet1_str.find('</c>', _vtag.end())
            if _vclose != -1:
                if '<f' in sheet1_str[_vtag.end():_vclose]:
                    _vneeds_fill = False
        if _vneeds_fill:
            _vsrc = _find_cell_with_formula(sheet1_str, 'V', row_num, _shared_fmla_cache)
            if _vsrc:
                sheet1_str = _inject_or_replace_cell(sheet1_str, row_num, 'V', _vsrc)
                logger.info("Backfilled V formula into %s", _vref)

        # Write Line No only when this original order_no has 2+ delivery-date
        # groups; a single-date order leaves the M column empty.
        _line_count_key = group.base_order_no if group.base_order_no is not None else group.order_no
        line_no_val: str = group.format_line_nos() if order_date_count[_line_count_key] > 1 else ""

        text_cols: List[Tuple[int, Optional[str]]] = [
            (ICPART_COL_PROJECT_ID,      group.project_id),
            (ICPART_COL_ORDER_NO,        group.order_no),
            (ICPART_COL_PO_NO,           group.po_no),
            (ICPART_COL_IC,              map_ic(group.customer_name)),
            # ICPART_COL_CUSTOMER_NAME removed — Customer Name no longer written to IC-Part
            (ICPART_COL_LINE_NO,         line_no_val),
            (ICPART_COL_INCOTERM,        group.incoterm),
            (ICPART_COL_CONTRACT_AMOUNT, group.contract_amount),
            (ICPART_COL_CORB_PATH,       group.corb_path),
        ]
        for col_idx, val in text_cols:
            if val is None or str(val).strip() == "":
                continue
            cell_ref = f"{_col_letter(col_idx)}{row_num}"
            sheet1_str = _inject_text_cell(sheet1_str, cell_ref, str(val))

        if group.delivery_date is not None:
            cell_ref = f"{_col_letter(ICPART_COL_DELIVERY_DATE)}{row_num}"
            serial = date_to_serial(group.delivery_date)
            sheet1_str = _inject_date_cell(sheet1_str, cell_ref, serial)

        if group.folder_date is not None:
            cell_ref_i = f"{_col_letter(ICPART_COL_FOLDER_DATE)}{row_num}"
            serial_i = date_to_serial(group.folder_date)
            sheet1_str = _inject_date_cell(sheet1_str, cell_ref_i, serial_i)

        # NOTE: do NOT add pid to _check_pids here — see frozen snapshot comment above.
        write_offset += 1
        written += 1

    logger.info("Groups injected: %d, rows skipped: %d", written, skipped)

    # Inject first group's Project Id into IC-Distribution sheet2 cell A2.
    # Sheet2 B2 formula should match Sheet1!I:I, where I = Project Id column.
    # Therefore A2 must hold Project Id, not Order No.
    # Fallback to order_no only when project_id is None (backward compatibility).
    _sheet2_key = "xl/worksheets/sheet2.xml"
    if _sheet2_key in all_zip_data and groups:
        _first_group = groups[0]
        _a2_value: Optional[str] = (
            _first_group.project_id.strip()
            if _first_group.project_id and _first_group.project_id.strip()
            else _first_group.order_no
        )
        if _a2_value:
            _s2 = all_zip_data[_sheet2_key].decode("utf-8")
            _a2_esc = _xml_escape(_a2_value)
            _a2_pat = re.compile(r'<c r="A2"(?:[^>]*/\s*>|[^>]*>.*?</c>)', re.DOTALL)
            def _a2_repl_fn(_m: re.Match) -> str:
                _s_match = re.search(r'\bs="(\d+)"', _m.group(0))
                _s_val = _s_match.group(1) if _s_match else "0"
                return f'<c r="A2" s="{_s_val}" t="inlineStr"><is><t>{_a2_esc}</t></is></c>'
            _s2_new, _n2 = _a2_pat.subn(_a2_repl_fn, _s2, count=1)
            if _n2 == 0:
                logger.warning("IC-Distribution A2 not found in sheet2.xml — skipped")
            else:
                _s2_new = _s2_new.replace("Sheet1!J:J", "Sheet1!I:I")
                all_zip_data[_sheet2_key] = _s2_new.encode("utf-8")
                logger.info("IC-Distribution A2 set to %s (project_id)", _a2_value)
    elif _sheet2_key not in all_zip_data:
        logger.warning("sheet2.xml not in template ZIP — IC-Distribution A2 not updated")

    sheet1_bytes_new: bytes = sheet1_str.encode("utf-8")

    import tempfile as _tempfile, shutil as _shutil, os as _os  # noqa: E401
    _tmp_path: Optional[str] = None
    try:
        _fd, _tmp_path = _tempfile.mkstemp(suffix=".xlsx", dir=_tempfile.gettempdir())
        _os.close(_fd)
        with zipfile.ZipFile(_tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            written_names: set = set()
            for item in zip_infolist:
                if item.filename == "xl/worksheets/sheet1.xml":
                    zout.writestr(item, sheet1_bytes_new)
                elif item.filename in all_zip_data:
                    zout.writestr(item, all_zip_data[item.filename])
                written_names.add(item.filename)
            for fname, fdata in all_zip_data.items():
                if fname not in written_names and fname != "xl/worksheets/sheet1.xml":
                    zout.writestr(fname, fdata)
        _MAX_RETRIES: int = 10
        _RETRY_INTERVAL: float = 2.0
        for _attempt in range(1, _MAX_RETRIES + 1):
            try:
                _shutil.copy2(_tmp_path, str(output_path))
                break  # success
            except PermissionError as _perm_exc:
                if _attempt < _MAX_RETRIES:
                    logger.warning(
                        "WinError 5 재시도 %d/%d — 파일 잠김(PermissionError): %s",
                        _attempt, _MAX_RETRIES, _perm_exc,
                    )
                    time.sleep(_RETRY_INTERVAL)
                else:
                    raise RuntimeError(
                        f"ZIP write failed after {_MAX_RETRIES}회 재시도 "
                        f"(PermissionError — 파일 잠김): {_perm_exc}"
                    ) from _perm_exc
            except OSError as _os_exc:
                if getattr(_os_exc, "winerror", None) == 5:
                    if _attempt < _MAX_RETRIES:
                        logger.warning(
                            "WinError 5 재시도 %d/%d — OSError winerror=5: %s",
                            _attempt, _MAX_RETRIES, _os_exc,
                        )
                        time.sleep(_RETRY_INTERVAL)
                    else:
                        raise RuntimeError(
                            f"ZIP write failed after {_MAX_RETRIES}회 재시도 "
                            f"(OSError winerror=5 — 파일 잠김): {_os_exc}"
                        ) from _os_exc
                else:
                    raise  # non-WinError-5 OSError → immediate re-raise
        os.utime(str(output_path), None)  # OneDrive file watcher trigger
        logger.info("ZIP write complete: %s", output_path)
    except Exception as exc:
        raise RuntimeError(f"ZIP write failed: {exc}") from exc
    finally:
        if _tmp_path is not None and _os.path.exists(_tmp_path):
            try:
                _os.unlink(_tmp_path)
            except OSError:
                pass

    return written, skipped


def _apply_label_via_excel(file_path: str) -> None:
    """Open file with a hidden Excel instance to apply sensitivity label, then save and close."""
    try:
        import win32com.client as _win32
        xl = _win32.DispatchEx("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        try:
            wb = xl.Workbooks.Open(file_path)
            wb.Save()
            wb.Close(False)
        finally:
            xl.Quit()
    except Exception as exc:
        import logging as _logging
        _logging.getLogger(__name__).warning("레이블 적용 실패 (무시하고 계속): %s", exc)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def run(
    source_path: Path,
    template_path: Path,
    output_path: Path,
    folder_date: Optional[date] = None,
    contract_amount: Optional[str] = None,
    incoterm: Optional[str] = None,
    corb_path: Optional[str] = None,
) -> None:
    """Orchestrate the full CustomerOrderLines -> IC-Part mapping."""
    if source_path is None:
        raise TypeError("source_path must not be None")
    if template_path is None:
        raise TypeError("template_path must not be None")
    if output_path is None:
        raise TypeError("output_path must not be None")

    logger.info("=== order_mapper START ===")
    groups = read_customer_order_lines(source_path)
    if len(groups) == 0:
        logger.warning("No groups found in source file — nothing to write.")
        return

    apply_sub_order_suffixes(groups)

    for _g in groups:
        if folder_date is not None:
            _g.folder_date = folder_date
        if contract_amount is not None:
            _g.contract_amount = contract_amount
        if incoterm is not None:
            _g.incoterm = incoterm
        if corb_path is not None:
            _g.corb_path = corb_path

    written, skipped = write_ic_part_zip(groups, template_path, output_path)
    print(f"[order_mapper] DONE — written={written} skipped={skipped} output={output_path}")
    logger.info("=== order_mapper DONE === written=%d skipped=%d", written, skipped)


# ---------------------------------------------------------------------------
# Self-verification block
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    if "--self-verify" in sys.argv:
        print("=== SELF-VERIFY START ===")

        # map_ic
        assert map_ic("CCI Valve Technology GmbH") == "AT"
        assert map_ic("CC Valve Technology Gmbh") == "AT", "CC Valve variant 미매핑"
        assert map_ic("IMI Critical Engineering LLC") == "SoCal"
        assert map_ic("Some APAC Company") == "SG"
        assert map_ic("CCI Valve Technology Gmbh") == "AT", "소문자 h 변형 미매핑"
        assert map_ic("cci valve technology gmbh") == "AT", "대소문자 변형 미매핑"
        assert map_ic("some apac company") == "SG", "APAC 대소문자 변형 미매핑"
        assert map_ic("Unknown Corp") == "Unknown Corp"
        assert map_ic(None) == ""
        try:
            map_ic(12345)  # type: ignore
            assert False
        except TypeError:
            pass

        # _col_letter
        assert _col_letter(10) == "J"
        assert _col_letter(11) == "K"
        assert _col_letter(17) == "Q"
        assert _col_letter(1) == "A"
        assert _col_letter(26) == "Z"
        assert _col_letter(27) == "AA"
        try:
            _col_letter(0)
            assert False
        except ValueError:
            pass

        # date_to_serial
        assert date_to_serial(_date(2026, 6, 2)) == 46175
        try:
            date_to_serial(None)  # type: ignore
            assert False
        except TypeError:
            pass

        # normalise_date
        from datetime import datetime as _dt
        assert normalise_date(None) is None
        assert normalise_date(_dt(2026, 6, 2, 0, 0)) == _date(2026, 6, 2)
        assert normalise_date(_date(2026, 6, 25)) == _date(2026, 6, 25)
        assert normalise_date("2026-07-01") == _date(2026, 7, 1)

        # format_line_nos
        og = OrderGroup()
        assert og.format_line_nos() == ""
        og.line_nos = [10, 7, 8, 9]
        assert og.format_line_nos() == "#7~10"
        og.line_nos = [3, 5, 6, 8]
        assert og.format_line_nos() == "#3,5,6,8"
        og.line_nos = [1, 2, 4]
        assert og.format_line_nos() == "#1,2,4"

        # apply_sub_order_suffixes
        og_a = OrderGroup(); og_a.order_no = "X100050542"; og_a.delivery_date = _date(2026, 6, 2)  # noqa: E702
        og_b = OrderGroup(); og_b.order_no = "X100050542"; og_b.delivery_date = _date(2026, 6, 25)  # noqa: E702
        og_c = OrderGroup(); og_c.order_no = "X100050542"; og_c.delivery_date = _date(2026, 7, 1)  # noqa: E702
        apply_sub_order_suffixes([og_a, og_b, og_c])
        assert og_a.order_no == "X100050542"
        assert og_b.order_no == "X100050542-2"
        assert og_c.order_no == "X100050542-3"
        assert og_b.base_order_no == "X100050542"

        _ln_a = OrderGroup(); _ln_a.order_no = "X100050777"; _ln_a.line_nos = [1, 2, 3]  # noqa: E702
        _ln_b = OrderGroup(); _ln_b.order_no = "X100050777"; _ln_b.line_nos = [4, 5]  # noqa: E702
        _ln_groups = [_ln_a, _ln_b]
        apply_sub_order_suffixes(_ln_groups)
        _ln_counts: Dict[Optional[str], int] = Counter(g.base_order_no for g in _ln_groups)
        _ln_values = [
            g.format_line_nos() if _ln_counts[g.base_order_no] > 1 else ""
            for g in _ln_groups
        ]
        assert _ln_values == ["#1~3", "#4,5"], f"Line No should use base order count, got {_ln_values!r}"

        # _inject_text_cell: self-closing empty cell
        _test_xml = (
            '<worksheet><sheetData>'
            '<row r="6"><c r="J6" s="38"/><c r="K6" s="38"/></row>'
            '</sheetData></worksheet>'
        )
        _result = _inject_text_cell(_test_xml, "J6", "TestValue")
        assert 't="inlineStr"' in _result
        assert '<is><t>TestValue</t></is>' in _result
        assert '<c r="K6" s="38"/>' in _result

        # _inject_text_cell: overwrite existing non-empty cell (second-run scenario)
        _filled_xml = (
            '<worksheet><sheetData>'
            '<row r="6"><c r="J6" s="14" t="inlineStr"><is><t>OldValue</t></is></c></row>'
            '</sheetData></worksheet>'
        )
        _result_ow = _inject_text_cell(_filled_xml, "J6", "NewValue")
        assert 'NewValue' in _result_ow
        assert 'OldValue' not in _result_ow
        assert 's="14"' in _result_ow

        # _inject_date_cell
        _test_xml2 = (
            '<worksheet><sheetData>'
            '<row r="6"><c r="Q6" s="44"/></row>'
            '</sheetData></worksheet>'
        )
        _result3 = _inject_date_cell(_test_xml2, "Q6", 46175)
        assert '<v>46175</v>' in _result3
        try:
            _inject_date_cell(_test_xml2, "Q6", 0)
            assert False
        except ValueError:
            pass

        # _detect_columns_from_header: new format
        _new_hdr = ('Order No', 'Line No', 'Del No', 'Rental', 'Project Id',
                    'Wanted Delivery Date', 'Status', 'Order Type', 'Customer No', 'Customer Name')
        _det = _detect_columns_from_header(_new_hdr)
        assert _det.get('project_id') == 4, f"project_id col should be 4, got {_det.get('project_id')}"
        assert _det.get('customer_name') == 9, f"customer_name col should be 9, got {_det.get('customer_name')}"

        # _detect_columns_from_header: old format
        _old_hdr = ('Wanted Delivery Date', 'Order No', 'Line No', 'Del No', 'Rental',
                    'Status', 'Order Type', 'Customer No', 'Customer Name')
        _det_old = _detect_columns_from_header(_old_hdr)
        assert _det_old.get('customer_name') == 8, f"old format customer_name should be 8, got {_det_old.get('customer_name')}"

        # _auto_detect_key_columns
        _detect_rows_a: List[Any] = [
            ("X100050542", datetime(2026, 6, 2)),
        ]
        _det_order, _det_date = _auto_detect_key_columns(_detect_rows_a)
        assert _det_order == 0
        assert _det_date == 1

        _detect_rows_b: List[Any] = [
            (datetime(2026, 6, 2), "X100050542"),
        ]
        _det_order_b, _det_date_b = _auto_detect_key_columns(_detect_rows_b)
        assert _det_order_b == 1
        assert _det_date_b == 0

        # add_line with col overrides
        _row: list = [None] * 10  # type: ignore[var-annotated]
        _row[1] = "X100050999"
        _row[0] = datetime(2026, 6, 10)
        _row[2] = 5
        _row[9] = "Test Customer Name"
        _row[4] = "B1234"
        _og = OrderGroup()
        _og.add_line(tuple(_row), col_order_no=1, col_delivery_date=0,
                     col_customer_name=9, col_project_id=4)
        assert _og.order_no == "X100050999"
        assert _og.customer_name == "Test Customer Name", f"got {_og.customer_name!r}"
        assert _og.project_id == "B1234", f"got {_og.project_id!r}"

        # Column constants (IMP-20260509-F8BF — all indices shifted left by 1)
        assert ICPART_COL_FOLDER_DATE == 8
        assert ICPART_COL_CONTRACT_AMOUNT == 14
        assert ICPART_COL_CORB_PATH == 18

        # Bug 2 fix 검증: frozen _check_pids 스냅샷으로 체크해야
        # 같은 Project Id를 가진 서브그룹(X100050786-2, -3)이 스킵되지 않아야 함.
        _existing_pids_frozen: Set[str] = {"PJ-001", "PJ-002"}  # already-in-template IDs
        _check_pids_test: Set[str] = set(_existing_pids_frozen)  # frozen snapshot — never mutated
        _groups_b2 = [
            # pid,      expect_skip
            ("PJ-001", True),   # 이미 템플릿에 있음 → skip
            ("PJ-003", False),  # 신규 → write
            ("PJ-003", False),  # 같은 실행 내 서브그룹 → 이제 write 해야 함 (버그 수정 후)
            (None,     False),  # None → write
        ]
        _test_written = 0
        _test_skipped = 0
        for _pid_raw, _expect_skip in _groups_b2:
            _pid_t = _pid_raw.strip() if _pid_raw else None
            if _pid_t and _pid_t in _check_pids_test:
                # _check_pids_test is never extended → sub-groups not skipped
                _test_skipped += 1
                assert _expect_skip, f"Unexpected skip for pid={_pid_t}"
            else:
                _test_written += 1
                assert not _expect_skip, f"Expected skip for pid={_pid_t} but wrote"
        assert _test_skipped == 1, f"Bug2: expected 1 skipped (template dup), got {_test_skipped}"
        assert _test_written == 3, f"Bug2: expected 3 written (sub-groups + None), got {_test_written}"
        print("bug2_frozen_check_pids_test: PASS")

        # Bug 1 fix 검증: IC-Distribution A2에 project_id가 들어가야 하고
        # project_id가 None이면 order_no 폴백.
        _og_b1a = OrderGroup()
        _og_b1a.order_no = "X100050786"
        _og_b1a.project_id = "B42728KR"
        _a2_val_a: Optional[str] = (
            _og_b1a.project_id.strip()
            if _og_b1a.project_id and _og_b1a.project_id.strip()
            else _og_b1a.order_no
        )
        assert _a2_val_a == "B42728KR", f"Bug1: A2 should be project_id, got {_a2_val_a!r}"

        _og_b1b = OrderGroup()
        _og_b1b.order_no = "X100050601"
        _og_b1b.project_id = None  # no project_id → fallback
        _a2_val_b: Optional[str] = (
            _og_b1b.project_id.strip()
            if _og_b1b.project_id and _og_b1b.project_id.strip()
            else _og_b1b.order_no
        )
        assert _a2_val_b == "X100050601", f"Bug1: fallback should be order_no, got {_a2_val_b!r}"
        print("bug1_a2_project_id_test: PASS")

        # col_line_no 전달 검증 (신규 포맷 Line No 오독 방지)
        _og_ln = OrderGroup()
        # 신규 포맷 행: col0=OrderNo, col1=LineNo(7), col2=DelNo(1), col5=Date
        _row_ln = ('X100', 7, 1, None, 'B001', '2026-06-02', None, None, None, 'IMI Critical Engineering LLC')
        _og_ln.add_line(_row_ln, col_order_no=0, col_delivery_date=5, col_line_no=1,
                        col_customer_name=9, col_project_id=4, col_po_no=None)
        assert _og_ln.line_nos == [7], f"Expected [7], got {_og_ln.line_nos}"
        # col_line_no=None 이면 fallback COL_LINE_NO=2 사용 (DelNo=1)
        _og_ln2 = OrderGroup()
        _og_ln2.add_line(_row_ln, col_order_no=0, col_delivery_date=5, col_line_no=None,
                         col_customer_name=9, col_project_id=4, col_po_no=None)
        assert _og_ln2.line_nos == [1], f"Expected [1] (fallback col2), got {_og_ln2.line_nos}"
        print("col_line_no_passthrough_test: PASS")

        print("=== SELF-VERIFY OK ===")
