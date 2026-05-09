"""
excel_handler.py — ExcelHandler class.

Reads a configurable cell range from an Excel workbook, marks processed cells
with a background colour, and checks whether the file was modified yesterday.
Uses openpyxl; no network I/O.
"""

import logging
import os
import re
import tempfile
from datetime import date, timedelta
from pathlib import Path
from typing import Any, List, Optional, Tuple

from core.config_manager import AppConfig, ExcelReadResult

logger = logging.getLogger(__name__)

# SEC-001: Allowed Excel file extensions
_ALLOWED_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}

# SEC-002: Hex colour validation
_HEX_COLOR_RE = re.compile(r'^[0-9A-Fa-f]{6}$')


def _resolve_excel_path(excel_path: Any) -> Optional[Path]:
    """Validate and resolve an Excel file path safely.

    Defends against path traversal, UNC paths, null bytes, oversized input,
    and non-Excel file extensions.

    Args:
        excel_path: The raw path value from config (expected str).

    Returns:
        A resolved Path if valid, or None if any validation check fails.
    """
    if not isinstance(excel_path, str) or not excel_path:
        return None
    if len(excel_path) > 260:
        logger.warning("_resolve_excel_path: path length > 260 — rejected")
        return None
    if "\x00" in excel_path:
        logger.warning("_resolve_excel_path: null byte detected — rejected")
        return None
    # Check raw input for UNC patterns before resolving.
    # Windows Path.resolve() may silently transform \\server\share into a local path.
    # A UNC path starts with exactly "\\" (two backslashes) or "//" (two forward slashes).
    if excel_path.startswith("\\\\") or excel_path.startswith("//"):
        logger.warning("_resolve_excel_path: UNC path rejected (raw input)")
        return None
    try:
        resolved = Path(excel_path).resolve()
    except (ValueError, OSError) as exc:
        logger.warning("_resolve_excel_path: resolve failed: %s", exc)
        return None
    resolved_str = str(resolved)
    if resolved_str.startswith("\\\\") or resolved_str.startswith("//"):
        logger.warning("_resolve_excel_path: UNC path rejected")
        return None
    if resolved.suffix.lower() not in _ALLOWED_EXTENSIONS:
        logger.warning(
            "_resolve_excel_path: disallowed extension '%s' — rejected",
            resolved.suffix,
        )
        return None
    return resolved


def _validate_hex_color(color: Any, fallback: str = "FFD700") -> str:
    """Validate and return a 6-digit hex colour string.

    Args:
        color:    The raw colour value to validate.
        fallback: The default colour to use when validation fails.

    Returns:
        The uppercased hex colour if valid, otherwise the fallback value.
    """
    if isinstance(color, str) and _HEX_COLOR_RE.match(color):
        return color.upper()
    logger.warning(
        "_validate_hex_color: invalid colour '%s' — using fallback '%s'", color, fallback
    )
    return fallback


class ExcelHandler:
    """Handles all Excel read / write operations for the automation pipeline."""

    def __init__(self, config: AppConfig) -> None:
        """Initialise with the application configuration.

        Args:
            config: The loaded AppConfig instance.
        """
        self._config = config

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def is_modified_yesterday(self) -> bool:
        """Return True if the Excel file's mtime date equals yesterday.

        EC-06: "yesterday" is defined as ``datetime.date.today() - timedelta(days=1)``.
        Returns False if the file does not exist or the mtime cannot be read.
        """
        excel_path = self._config.get("excel_path", "")
        if not excel_path:
            logger.warning("is_modified_yesterday: excel_path is empty — returning False")
            return False

        target_path = _resolve_excel_path(excel_path)
        if target_path is None:
            logger.warning("is_modified_yesterday: invalid excel_path — returning False")
            return False

        if not target_path.is_file():
            logger.warning(
                "is_modified_yesterday: file not found at %s — returning False",
                target_path,
            )
            return False

        try:
            mtime_ts = target_path.stat().st_mtime
            mtime_date = date.fromtimestamp(mtime_ts)
        except OSError as exc:
            logger.error("is_modified_yesterday: cannot stat %s: %s", target_path, exc)
            return False

        yesterday = date.today() - timedelta(days=1)
        result = mtime_date == yesterday
        logger.debug(
            "is_modified_yesterday: file mtime date=%s, yesterday=%s → %s",
            mtime_date,
            yesterday,
            result,
        )
        return result

    def read_cells(self) -> ExcelReadResult:
        """Read the configured cell range from the Excel workbook.

        EC-02: Returns success=False (no exception propagated) if the file is
        missing or locked.
        EC-03: Returns success=True with empty data_as_text if the range
        contains no non-None values.

        Returns:
            ExcelReadResult with success flag, 2-D data list, plain-text
            representation and an optional error string.
        """
        excel_path = self._config.get("excel_path", "")
        if not excel_path:
            logger.error("read_cells: excel_path is empty")
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="excel_path is not configured",
            )

        target_path = _resolve_excel_path(excel_path)
        if target_path is None:
            logger.error("read_cells: invalid or unsafe excel_path")
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="Excel 파일 경로가 유효하지 않습니다.",
            )

        if not target_path.is_file():
            logger.error("read_cells: file not found — path=%s", target_path)
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="Excel 파일을 찾을 수 없습니다.",
            )

        cell_range = self._config.get("cell_range", {})
        sheet_name: str = cell_range.get("sheet_name", "Sheet1")
        start_row: int = cell_range.get("start_row", 1)
        start_col: int = cell_range.get("start_col", 1)
        end_row: int = cell_range.get("end_row", 10)
        end_col: int = cell_range.get("end_col", 5)

        # Validate range bounds
        if start_row > end_row or start_col > end_col:
            logger.error(
                "read_cells: invalid range (rows %d-%d, cols %d-%d)",
                start_row, end_row, start_col, end_col,
            )
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="Invalid cell range: start exceeds end",
            )

        try:
            import openpyxl  # type: ignore[import]
        except ImportError as exc:
            logger.error("read_cells: openpyxl not installed: %s", exc)
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="openpyxl is not installed",
            )

        try:
            workbook = openpyxl.load_workbook(str(target_path), read_only=False, data_only=True)
        except PermissionError as exc:
            logger.error("read_cells: file is locked: %s — path=%s", exc, target_path)
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="Excel 파일에 접근할 수 없습니다. 파일이 열려 있거나 권한이 없습니다.",
            )
        except Exception as exc:
            logger.error("read_cells: cannot open workbook: %s — path=%s", exc, target_path)
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="Excel 파일을 열 수 없습니다.",
            )

        try:
            if sheet_name not in workbook.sheetnames:
                logger.warning(
                    "read_cells: sheet '%s' not found; available: %s",
                    sheet_name,
                    workbook.sheetnames,
                )
                return ExcelReadResult(
                    success=False,
                    data=[],
                    data_as_text="",
                    error_msg="지정한 시트를 찾을 수 없습니다.",
                )

            ws = workbook[sheet_name]
            rows_data: List[List[Any]] = []

            for row_idx in range(start_row, end_row + 1):
                row_cells: List[Any] = []
                for col_idx in range(start_col, end_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    row_cells.append(cell.value)
                rows_data.append(row_cells)

        except Exception as exc:
            logger.error("read_cells: error reading cells: %s — path=%s", exc, target_path)
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="셀 데이터를 읽는 중 오류가 발생했습니다.",
            )
        finally:
            try:
                workbook.close()
            except Exception:
                pass

        # EC-03: check for empty range
        text = self._cells_to_text(rows_data)
        logger.info(
            "read_cells: read %d rows × %d cols from '%s'",
            len(rows_data),
            end_col - start_col + 1,
            sheet_name,
        )
        return ExcelReadResult(success=True, data=rows_data, data_as_text=text, error_msg=None)

    def mark_cells(self) -> bool:
        """Apply the configured mark_color as a cell background to the range.

        Saves the workbook after applying fills.
        Returns True on success, False on any failure (caller continues
        regardless — pipeline treats mark failure as non-fatal).
        """
        excel_path = self._config.get("excel_path", "")
        if not excel_path:
            logger.warning("mark_cells: excel_path is empty — skipping")
            return False

        target_path = _resolve_excel_path(excel_path)
        if target_path is None:
            logger.error("mark_cells: invalid or unsafe excel_path — skipping")
            return False

        if not target_path.is_file():
            logger.error("mark_cells: file not found — skipping")
            return False

        raw_color: str = self._config.get("mark_color", "FFD700")
        mark_color: str = _validate_hex_color(raw_color)

        cell_range = self._config.get("cell_range", {})
        sheet_name: str = cell_range.get("sheet_name", "Sheet1")
        start_row: int = cell_range.get("start_row", 1)
        start_col: int = cell_range.get("start_col", 1)
        end_row: int = cell_range.get("end_row", 10)
        end_col: int = cell_range.get("end_col", 5)

        try:
            import openpyxl  # type: ignore[import]
            from openpyxl.styles import PatternFill  # type: ignore[import]
        except ImportError as exc:
            logger.error("mark_cells: openpyxl not installed: %s", exc)
            return False

        try:
            workbook = openpyxl.load_workbook(str(target_path), read_only=False, data_only=False)
        except PermissionError as exc:
            logger.error("mark_cells: file is locked: %s", exc)
            return False
        except Exception as exc:
            logger.error("mark_cells: cannot open workbook: %s", exc)
            return False

        try:
            if sheet_name not in workbook.sheetnames:
                logger.error("mark_cells: sheet '%s' not found", sheet_name)
                return False

            ws = workbook[sheet_name]
            fill = PatternFill(
                start_color=mark_color,
                end_color=mark_color,
                fill_type="solid",
            )
            for row_idx in range(start_row, end_row + 1):
                for col_idx in range(start_col, end_col + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

            # FS.safe_write: atomic save via tempfile → os.replace
            tmp_fd, tmp_path = tempfile.mkstemp(
                suffix=".xlsx", dir=str(target_path.parent)
            )
            try:
                os.close(tmp_fd)
                workbook.save(tmp_path)
                os.replace(tmp_path, str(target_path))
            except Exception:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
                raise
            logger.info("mark_cells: applied fill %s to range and saved atomically", mark_color)
            return True

        except Exception as exc:
            logger.error("mark_cells: error applying fill: %s", exc)
            return False
        finally:
            try:
                workbook.close()
            except Exception:
                pass

    def detect_end_row(
        self,
        ws: Any,
        start_row: int,
        col_start: int,
        col_end: int,
    ) -> int:
        """Scan downward from start_row and return the last non-empty row index.

        "Non-empty" means at least one cell in col_start..col_end has a value
        that is not None and not an empty string.

        Returns:
            The last non-empty row number (≥ start_row), or 0 when
            start_row itself is already completely empty (no data to process).

        Args:
            ws:        An openpyxl Worksheet object.
            start_row: The first row index to examine (1-indexed, must be ≥ 1).
            col_start: The first column index (1-indexed, must be ≥ 1).
            col_end:   The last  column index (1-indexed, must be ≥ col_start).

        Raises:
            TypeError:  If any parameter is None or not an int.
            ValueError: If start_row, col_start, or col_end are ≤ 0, or
                        col_end < col_start.
        """
        # AL.type_valid — None + isinstance + boundary for each int param
        for param_name, param_val in (
            ("ws", ws),
        ):
            if param_val is None:
                raise TypeError(f"{param_name} must not be None")

        if start_row is None:
            raise TypeError("start_row must not be None")
        if isinstance(start_row, bool) or not isinstance(start_row, int):
            raise TypeError(f"start_row must be int, got {type(start_row).__name__}")
        # positive integer allowed — 1-indexed Excel row number
        if start_row <= 0:
            raise ValueError(f"start_row must be positive, got {start_row}")

        if col_start is None:
            raise TypeError("col_start must not be None")
        if isinstance(col_start, bool) or not isinstance(col_start, int):
            raise TypeError(f"col_start must be int, got {type(col_start).__name__}")
        # positive integer allowed — 1-indexed Excel column number
        if col_start <= 0:
            raise ValueError(f"col_start must be positive, got {col_start}")

        if col_end is None:
            raise TypeError("col_end must not be None")
        if isinstance(col_end, bool) or not isinstance(col_end, int):
            raise TypeError(f"col_end must be int, got {type(col_end).__name__}")
        # positive integer allowed — 1-indexed Excel column number
        if col_end <= 0:
            raise ValueError(f"col_end must be positive, got {col_end}")

        if col_end < col_start:
            raise ValueError(
                f"col_end ({col_end}) must be >= col_start ({col_start})"
            )

        def _row_is_empty(row_idx: int) -> bool:
            """Return True when every cell in col_start..col_end is None or ''."""
            for col_idx in range(col_start, col_end + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is not None and str(val).strip() != "":
                    return False
            return True

        # If the very first row is empty, signal "nothing to process"
        if _row_is_empty(start_row):
            logger.debug(
                "detect_end_row: start_row=%d is already empty → returning 0",
                start_row,
            )
            return 0

        last_non_empty = start_row
        row_idx = start_row + 1

        # Walk down until we hit an empty row or exceed max worksheet dimension
        max_row: int = ws.max_row or start_row
        while row_idx <= max_row:
            if _row_is_empty(row_idx):
                break
            last_non_empty = row_idx
            row_idx += 1

        logger.debug(
            "detect_end_row: start_row=%d → end_row=%d",
            start_row,
            last_non_empty,
        )
        return last_non_empty

    def read_cells_dynamic(
        self,
        start_row: int,
        end_row: int,
        col_start: int,
        col_end: int,
    ) -> ExcelReadResult:
        """Read a dynamically supplied cell range from the configured workbook.

        Accepts an explicit row/column range rather than reading it from config,
        enabling the caller to pass detected boundaries from detect_end_row().

        Args:
            start_row: First row to read (1-indexed, must be ≥ 1).
            end_row:   Last  row to read (1-indexed, must be ≥ start_row).
            col_start: First column to read (1-indexed, must be ≥ 1).
            col_end:   Last  column to read (1-indexed, must be ≥ col_start).

        Returns:
            ExcelReadResult with success flag, 2-D data list, plain-text
            representation and an optional error string.

        Raises:
            TypeError:  If any parameter is None or not an int.
            ValueError: If range boundaries are invalid (≤ 0 or start > end).
        """
        # AL.type_valid for all four int parameters
        for param_name, param_val in (
            ("start_row", start_row),
            ("end_row", end_row),
            ("col_start", col_start),
            ("col_end", col_end),
        ):
            if param_val is None:
                raise TypeError(f"{param_name} must not be None")
            if isinstance(param_val, bool) or not isinstance(param_val, int):
                raise TypeError(
                    f"{param_name} must be int, got {type(param_val).__name__}"
                )
            # positive integer allowed — 1-indexed Excel row/column numbers
            if param_val <= 0:
                raise ValueError(f"{param_name} must be positive, got {param_val}")

        if start_row > end_row:
            raise ValueError(
                f"start_row ({start_row}) must be <= end_row ({end_row})"
            )
        if col_start > col_end:
            raise ValueError(
                f"col_start ({col_start}) must be <= col_end ({col_end})"
            )

        excel_path = self._config.get("excel_path", "")
        if not excel_path:
            logger.error("read_cells_dynamic: excel_path is empty")
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="excel_path is not configured",
            )

        target_path = _resolve_excel_path(excel_path)
        if target_path is None:
            logger.error("read_cells_dynamic: invalid or unsafe excel_path")
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="Excel 파일 경로가 유효하지 않습니다.",
            )

        if not target_path.is_file():
            logger.error("read_cells_dynamic: file not found — path=%s", target_path)
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="Excel 파일을 찾을 수 없습니다.",
            )

        cell_range = self._config.get("cell_range", {})
        sheet_name: str = cell_range.get("sheet_name", "Sheet1")

        try:
            import openpyxl  # type: ignore[import]
        except ImportError as exc:
            logger.error("read_cells_dynamic: openpyxl not installed: %s", exc)
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="openpyxl is not installed",
            )

        try:
            workbook = openpyxl.load_workbook(
                str(target_path), read_only=False, data_only=True
            )
        except PermissionError as exc:
            logger.error("read_cells_dynamic: file locked: %s", exc)
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="Excel 파일에 접근할 수 없습니다.",
            )
        except Exception as exc:
            logger.error("read_cells_dynamic: cannot open workbook: %s", exc)
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="Excel 파일을 열 수 없습니다.",
            )

        try:
            if sheet_name not in workbook.sheetnames:
                logger.warning(
                    "read_cells_dynamic: sheet '%s' not found; available: %s",
                    sheet_name,
                    workbook.sheetnames,
                )
                return ExcelReadResult(
                    success=False,
                    data=[],
                    data_as_text="",
                    error_msg="지정한 시트를 찾을 수 없습니다.",
                )

            ws = workbook[sheet_name]
            rows_data: List[List[Any]] = []

            for row_idx in range(start_row, end_row + 1):
                row_cells: List[Any] = []
                for col_idx in range(col_start, col_end + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    row_cells.append(cell.value)
                rows_data.append(row_cells)

        except Exception as exc:
            logger.error("read_cells_dynamic: error reading cells: %s", exc)
            return ExcelReadResult(
                success=False,
                data=[],
                data_as_text="",
                error_msg="셀 데이터를 읽는 중 오류가 발생했습니다.",
            )
        finally:
            try:
                workbook.close()
            except Exception:
                pass

        text = self._cells_to_text(rows_data)
        logger.info(
            "read_cells_dynamic: read %d rows × %d cols from '%s'",
            len(rows_data),
            col_end - col_start + 1,
            sheet_name,
        )
        return ExcelReadResult(
            success=True, data=rows_data, data_as_text=text, error_msg=None
        )

    def rows_to_html_table(
        self,
        ws: Any,
        start_row: int,
        end_row: int,
        col_start: int,
        col_end: int,
    ) -> str:
        """Convert a worksheet range to an HTML table string.

        Inline styles are applied for bold (font-weight:bold) and fill colour
        (background-color:#RRGGBB).  MergedCell instances are rendered as
        empty <td> elements (span handling is simplified per spec).

        Args:
            ws:        An openpyxl Worksheet object.
            start_row: First row to render (1-indexed, must be ≥ 1).
            end_row:   Last  row to render (1-indexed, must be ≥ start_row).
            col_start: First column (1-indexed, must be ≥ 1).
            col_end:   Last  column (1-indexed, must be ≥ col_start).

        Returns:
            An HTML string beginning with <table> and ending with </table>.

        Raises:
            TypeError:  If ws is None, or any int parameter is None/non-int.
            ValueError: If range boundaries are invalid.
        """
        # AL.type_valid for ws
        if ws is None:
            raise TypeError("ws must not be None")

        # AL.type_valid for int params
        for param_name, param_val in (
            ("start_row", start_row),
            ("end_row", end_row),
            ("col_start", col_start),
            ("col_end", col_end),
        ):
            if param_val is None:
                raise TypeError(f"{param_name} must not be None")
            if isinstance(param_val, bool) or not isinstance(param_val, int):
                raise TypeError(
                    f"{param_name} must be int, got {type(param_val).__name__}"
                )
            # positive integer allowed — 1-indexed Excel row/column numbers
            if param_val <= 0:
                raise ValueError(f"{param_name} must be positive, got {param_val}")

        if start_row > end_row:
            raise ValueError(
                f"start_row ({start_row}) must be <= end_row ({end_row})"
            )
        if col_start > col_end:
            raise ValueError(
                f"col_start ({col_start}) must be <= col_end ({col_end})"
            )

        try:
            from openpyxl.cell import MergedCell  # type: ignore[import]
        except ImportError:
            # Fallback: treat no cell as a MergedCell
            class MergedCell:  # type: ignore[no-redef]
                pass

        rows_html: List[str] = []
        for row_idx in range(start_row, end_row + 1):
            cells_html: List[str] = []
            for col_idx in range(col_start, col_end + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # MergedCell: render as empty td (simplified — no span)
                if isinstance(cell, MergedCell):
                    cells_html.append("<td></td>")
                    continue

                styles: List[str] = []

                # Bold detection
                try:
                    if cell.font and cell.font.bold:
                        styles.append("font-weight:bold")
                except Exception:
                    pass

                # Fill colour detection
                try:
                    fill = cell.fill
                    if (
                        fill
                        and fill.fill_type == "solid"
                        and fill.fgColor
                        and fill.fgColor.type == "rgb"
                    ):
                        rgb_val: str = fill.fgColor.rgb  # e.g. "FFFF0000"
                        # Strip alpha channel if present (first 2 hex digits)
                        if len(rgb_val) == 8:
                            rgb_val = rgb_val[2:]
                        if _HEX_COLOR_RE.match(rgb_val):
                            styles.append(f"background-color:#{rgb_val}")
                except Exception:
                    pass

                style_attr = f' style="{";".join(styles)}"' if styles else ""
                cell_value: str = (
                    "" if cell.value is None else str(cell.value)
                )
                cells_html.append(f"<td{style_attr}>{cell_value}</td>")

            rows_html.append("<tr>" + "".join(cells_html) + "</tr>")

        table_style = (
            'style="border-collapse:collapse;font-family:Arial,sans-serif;'
            'font-size:12px;"'
        )
        html = f"<table {table_style}>" + "".join(rows_html) + "</table>"
        logger.debug(
            "rows_to_html_table: rendered %d rows × %d cols",
            end_row - start_row + 1,
            col_end - col_start + 1,
        )
        return html

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _cells_to_text(self, data: List[List[Any]]) -> str:
        """Convert a 2-D cell array to a tab + newline delimited string.

        None values are converted to empty strings.

        Args:
            data: 2-D list where each inner list is one row of cell values.

        Returns:
            A string with cells separated by tabs and rows separated by newlines.
            Returns an empty string if data is empty or all values are None.
        """
        if not data:
            return ""

        lines: List[str] = []
        for row in data:
            cells: List[str] = []
            for cell in row:
                cells.append("" if cell is None else str(cell))
            lines.append("\t".join(cells))

        result = "\n".join(lines)
        # If every cell was None the result is a block of tabs/newlines; treat as empty.
        if not result.strip():
            return ""
        return result
