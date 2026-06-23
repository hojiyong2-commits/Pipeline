"""Code-based acceptance scorers for pipeline v2."""

from __future__ import annotations

import csv
import json
import os
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any, Callable


class ScoringError(Exception):
    """Raised when a scorer cannot evaluate its input."""


def _resolve(base_dir: Path, raw_path: str | None) -> Path:
    if not raw_path:
        raise ScoringError("path is required")
    path = Path(raw_path)
    if not path.is_absolute():
        path = base_dir / path
    return path


def _smart_resolve(base_dir: str, path_str: str, project_dir: str | None = None) -> str:
    """상대 경로를 프로젝트 루트 우선으로 해석한다.

    절대 경로는 그대로 반환한다.
    상대 경로는 project_dir이 주어지면 그것을 먼저 탐색하고,
    없으면 프로젝트 루트(scorers.py의 상위 세 디렉터리)를 탐색하고,
    없으면 base_dir 기준으로 해석한다.

    scorers.py 파일 위치: <project_root>/core/acceptance/scorers.py
    따라서 parent.parent.parent가 실제 project_root이다.
    (parent = acceptance/, parent.parent = core/, parent.parent.parent = project root)
    """
    p = Path(path_str)
    if p.is_absolute():
        return str(p)
    # project_dir이 명시적으로 전달된 경우 최우선 탐색
    if project_dir is not None:
        candidate = Path(project_dir) / path_str
        if candidate.exists():
            return str(candidate)
    # scorers.py 기준 project_root 탐색 (parent.parent.parent)
    project_root = Path(__file__).resolve().parent.parent.parent
    candidate = project_root / path_str
    if candidate.exists():
        return str(candidate)
    return str(Path(base_dir) / path_str)


def _load_json(path: Path) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except FileNotFoundError as exc:
        raise ScoringError(f"missing JSON file: {path}") from exc
    except json.JSONDecodeError as exc:
        raise ScoringError(f"invalid JSON in {path}: {exc}") from exc


def _normalize(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, list):
        return [_normalize(item) for item in value]
    if isinstance(value, dict):
        return {str(k): _normalize(v) for k, v in sorted(value.items())}
    return value


def _expected_actual(test: dict[str, Any], base_dir: Path, project_dir: Path | None = None) -> tuple[Any, Any]:
    then = test.get("then", {})
    given = test.get("given", {})
    expected = then.get("expected") if isinstance(then, dict) else None
    actual = given.get("actual") if isinstance(given, dict) else None

    # Resolve file paths: prefer project_dir for relative paths that start with
    # known project-root subdirectories (tests/, src/, etc.) so that test_set.json
    # entries that use project-root relative paths work correctly even when
    # base_dir is the contract directory.
    # NOTE: named _resolve_path (not _smart_resolve) to avoid shadowing the
    # module-level _smart_resolve(base_dir, path_str) helper.
    def _resolve_path(raw: str) -> Path:
        p = Path(raw)
        if p.is_absolute():
            return p
        # Try project_dir first (project-root relative path)
        if project_dir is not None:
            candidate = project_dir / p
            if candidate.exists():
                return candidate
        # Fall back to base_dir (contract-dir relative)
        return base_dir / p

    if isinstance(then, dict) and then.get("expected_file"):
        expected = _load_json(_resolve_path(str(then["expected_file"])))
    if isinstance(given, dict) and given.get("actual_file"):
        actual = _load_json(_resolve_path(str(given["actual_file"])))
    return expected, actual


def score_json_exact_match(test: dict[str, Any], base_dir: Path, project_dir: Path) -> tuple[bool, str, dict[str, Any]]:
    expected, actual = _expected_actual(test, base_dir, project_dir)
    ok = _normalize(expected) == _normalize(actual)
    return ok, "JSON values match" if ok else "JSON values differ", {"expected": expected, "actual": actual}


def score_mapping_rule_check(test: dict[str, Any], base_dir: Path, project_dir: Path) -> tuple[bool, str, dict[str, Any]]:
    return score_json_exact_match(test, base_dir, project_dir)


def score_email_parse_check(test: dict[str, Any], base_dir: Path, project_dir: Path) -> tuple[bool, str, dict[str, Any]]:
    return score_json_exact_match(test, base_dir, project_dir)


def score_file_exists_check(test: dict[str, Any], base_dir: Path, project_dir: Path) -> tuple[bool, str, dict[str, Any]]:
    then = test.get("then", {})
    path_value = then.get("path") if isinstance(then, dict) else None
    resolved = _smart_resolve(str(base_dir), str(path_value) if path_value else "", str(project_dir))
    path = Path(resolved) if resolved else _resolve(base_dir, str(path_value) if path_value else None)
    ok = path.exists()
    return ok, f"file exists: {path}" if ok else f"file missing: {path}", {"path": str(path)}


def score_file_output_check(test: dict[str, Any], base_dir: Path, project_dir: Path) -> tuple[bool, str, dict[str, Any]]:
    then = test.get("then", {})
    if not isinstance(then, dict):
        raise ScoringError("then object is required")
    raw_path = str(then.get("path") or "")
    path = Path(_smart_resolve(str(base_dir), raw_path, str(project_dir))) if raw_path else _resolve(base_dir, raw_path)
    if not path.exists():
        return False, f"file missing: {path}", {"path": str(path)}
    contains = then.get("contains")
    if contains is not None:
        text = path.read_text(encoding="utf-8", errors="replace")
        # contains가 리스트이면 모든 항목이 파일에 있는지 확인한다.
        if isinstance(contains, list):
            missing = [s for s in contains if str(s) not in text]
            ok = not missing
            return ok, "file contains expected text" if ok else "file does not contain expected text", {
                "path": str(path),
                "contains": contains,
            }
        ok = str(contains) in text
        return ok, "file contains expected text" if ok else "file does not contain expected text", {
            "path": str(path),
            "contains": contains,
        }
    return True, f"file exists: {path}", {"path": str(path)}


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def score_csv_row_match(test: dict[str, Any], base_dir: Path, project_dir: Path) -> tuple[bool, str, dict[str, Any]]:
    given = test.get("given", {})
    then = test.get("then", {})
    if not isinstance(given, dict) or not isinstance(then, dict):
        raise ScoringError("given and then objects are required")
    actual_path = _resolve(base_dir, str(given.get("actual_file") or given.get("file") or ""))
    expected_rows = then.get("expected_rows")
    if expected_rows is None and then.get("expected_file"):
        expected_rows = _read_csv_rows(_resolve(base_dir, str(then["expected_file"])))
    if not isinstance(expected_rows, list):
        raise ScoringError("then.expected_rows or then.expected_file is required")
    actual_rows = _read_csv_rows(actual_path)
    ok = _normalize(actual_rows) == _normalize(expected_rows)
    return ok, "CSV rows match" if ok else "CSV rows differ", {
        "actual_file": str(actual_path),
        "expected_rows": expected_rows,
        "actual_rows": actual_rows,
    }


def _load_workbook(path: Path) -> Any:
    try:
        import openpyxl  # type: ignore[import]
    except ImportError as exc:
        raise ScoringError("openpyxl is required for Excel acceptance tests") from exc
    try:
        return openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        raise ScoringError(f"failed to open workbook {path}: {exc}") from exc


def _worksheet_rows(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    workbook = _load_workbook(path)
    try:
        ws = workbook[sheet_name] if sheet_name else workbook.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        result = []
        for raw_row in rows[1:]:
            item = {headers[i]: raw_row[i] for i in range(min(len(headers), len(raw_row))) if headers[i]}
            if any(value is not None and value != "" for value in item.values()):
                result.append(item)
        return result
    finally:
        workbook.close()


def score_excel_schema_check(test: dict[str, Any], base_dir: Path, project_dir: Path) -> tuple[bool, str, dict[str, Any]]:
    given = test.get("given", {})
    then = test.get("then", {})
    if not isinstance(given, dict) or not isinstance(then, dict):
        raise ScoringError("given and then objects are required")
    path = _resolve(base_dir, str(given.get("file") or given.get("actual_file") or ""))
    rows = _worksheet_rows(path, given.get("sheet"))
    headers = set(rows[0].keys()) if rows else set()
    expected_columns = [str(col) for col in then.get("expected_columns", [])]
    missing = [col for col in expected_columns if col not in headers]
    ok = not missing
    return ok, "Excel schema matches" if ok else f"missing Excel columns: {missing}", {
        "file": str(path),
        "expected_columns": expected_columns,
        "actual_columns": sorted(headers),
        "missing": missing,
    }


def score_excel_row_match(test: dict[str, Any], base_dir: Path, project_dir: Path) -> tuple[bool, str, dict[str, Any]]:
    given = test.get("given", {})
    then = test.get("then", {})
    if not isinstance(given, dict) or not isinstance(then, dict):
        raise ScoringError("given and then objects are required")
    actual_path = _resolve(base_dir, str(given.get("actual_file") or given.get("file") or ""))
    actual_rows = _worksheet_rows(actual_path, given.get("sheet"))
    expected_rows = then.get("expected_rows")
    if expected_rows is None and then.get("expected_file"):
        expected_rows = _worksheet_rows(_resolve(base_dir, str(then["expected_file"])), then.get("sheet"))
    if not isinstance(expected_rows, list):
        raise ScoringError("then.expected_rows or then.expected_file is required")
    ok = _normalize(actual_rows) == _normalize(expected_rows)
    return ok, "Excel rows match" if ok else "Excel rows differ", {
        "actual_file": str(actual_path),
        "expected_rows": expected_rows,
        "actual_rows": actual_rows,
    }


def score_exe_launch_check(test: dict[str, Any], base_dir: Path, project_dir: Path) -> tuple[bool, str, dict[str, Any]]:
    given = test.get("given", {})
    if not isinstance(given, dict):
        raise ScoringError("given object is required")
    exe_path = _resolve(project_dir, str(given.get("exe") or given.get("path") or ""))
    if not exe_path.exists():
        return False, f"EXE missing: {exe_path}", {"exe": str(exe_path)}
    timeout = int(given.get("timeout_seconds", 5))
    args = given.get("args", [])
    if not isinstance(args, list):
        raise ScoringError("given.args must be a list")
    try:
        proc = subprocess.Popen(
            [str(exe_path), *[str(arg) for arg in args]],
            cwd=str(exe_path.parent),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        try:
            stdout, stderr = proc.communicate(timeout=timeout)
            ok = proc.returncode == int(given.get("expected_returncode", 0))
            return ok, "EXE exited with expected code" if ok else "EXE returned unexpected code", {
                "exe": str(exe_path),
                "returncode": proc.returncode,
                "stdout": stdout[-2000:],
                "stderr": stderr[-2000:],
            }
        except subprocess.TimeoutExpired:
            proc.terminate()
            return True, "EXE launched and stayed alive until smoke timeout", {"exe": str(exe_path)}
    except OSError as exc:
        return False, f"EXE launch failed: {exc}", {"exe": str(exe_path)}


def _build_command_check_env(
    given: dict[str, Any],
    base_dir: Path,
    project_dir: Path,
    when: dict[str, Any] | None = None,
) -> tuple[dict[str, str], str | None]:
    """given/when 블록의 isolation/fixture/env를 처리하여 subprocess 환경변수를 반환한다.

    Returns:
        (env_dict, tmp_dir_to_cleanup) — tmp_dir_to_cleanup은 사용 후 삭제할 임시 디렉토리
        경로이거나 None.
    """
    if when is None:
        when = {}
    env = dict(os.environ)
    tmp_dir: str | None = None

    isolation = given.get("isolation")
    # IMP-20260623-7EAA Round 2: given.cleanup_status=BLOCKED → 자동 격리 state로 사전 조건 설정.
    # test_set.json이 frozen이라 isolation 필드를 추가할 수 없는 oracle TC 를 위한 범용 fallback.
    # pipeline type 은 given 에서 읽어 IMP 전용 하드코딩을 제거한다(범용화). 미지정 시 기본 IMP.
    cleanup_status = given.get("cleanup_status")
    if cleanup_status == "BLOCKED" and isolation is None:
        pipeline_type = given.get("pipeline_type", "IMP")  # 범용 — given에서 type 읽기 허용
        # post_complete_cleanup.status=BLOCKED인 격리 state 생성
        state_content: dict[str, Any] = {
            "version": "1.2.0",
            "pipeline_id": "ORACLE-TC-SETUP",
            "type": pipeline_type,
            "description": "oracle TC precondition: cleanup BLOCKED",
            "created_at": "2026-01-01T00:00:00Z",
            "updated_at": "2026-01-01T00:00:00Z",
            "current_phase": "architect",
            "blocked": False,
            "blocked_reason": None,
            "post_complete_cleanup": {
                "status": "BLOCKED",
                "ready_for_next_task": False,
                "removed": [],
                "preserved": [],
                "deferred": [],
                "missing_protected": ["oracle_tc5_precondition.json"],
                "completed_at": "2026-01-01T00:00:00Z",
                "reason": "oracle TC-5 precondition: previous pipeline cleanup BLOCKED",
            },
        }
        tmp_dir = tempfile.mkdtemp(prefix="oracle_cmd_check_")
        tmp_state = os.path.join(tmp_dir, "oracle_isolated_state.json")
        with open(tmp_state, "w", encoding="utf-8") as f:
            json.dump(state_content, f, ensure_ascii=False, indent=2)
        env["PIPELINE_STATE_PATH"] = tmp_state
    elif isolation == "PIPELINE_STATE_PATH":
        fixture_path = given.get("fixture")
        if fixture_path:
            # fixture 파일에서 state를 읽어 임시 파일에 씀
            resolved_fixture = Path(fixture_path)
            if not resolved_fixture.is_absolute():
                resolved_fixture = project_dir / fixture_path
            try:
                fixture_data = json.loads(resolved_fixture.read_text(encoding="utf-8"))
                state_content = fixture_data.get("preconditions", {}).get("state_file") or {}
            except (OSError, json.JSONDecodeError, AttributeError):
                state_content = {}
        else:
            state_content = {}
        tmp_dir = tempfile.mkdtemp(prefix="oracle_cmd_check_")
        tmp_state = os.path.join(tmp_dir, "oracle_isolated_state.json")
        with open(tmp_state, "w", encoding="utf-8") as f:
            json.dump(state_content, f, ensure_ascii=False, indent=2)
        env["PIPELINE_STATE_PATH"] = tmp_state

    # given.env 오버라이드 지원
    extra_env = given.get("env")
    if isinstance(extra_env, dict):
        env.update({str(k): str(v) for k, v in extra_env.items()})

    # when.env 오버라이드 지원 (given.env보다 낮은 우선순위 — given이 더 구체적)
    # test_set.json에서 when.env에 환경변수를 지정하는 경우를 지원한다.
    when_env = when.get("env")
    if isinstance(when_env, dict):
        for k, v in when_env.items():
            k_str = str(k)
            # given.env가 이미 같은 키를 설정했다면 when.env는 덮어쓰지 않는다.
            if k_str not in (extra_env or {}):
                env[k_str] = str(v)

    return env, tmp_dir


def score_command_check(test: dict[str, Any], base_dir: Path, project_dir: Path) -> tuple[bool, str, dict[str, Any]]:
    when = test.get("when", {})
    then = test.get("then", {})
    given = test.get("given", {})
    if not isinstance(when, dict) or not isinstance(then, dict):
        raise ScoringError("when and then objects are required")
    if not isinstance(given, dict):
        given = {}
    command = when.get("command")
    if not isinstance(command, list) or not command:
        raise ScoringError("when.command must be a non-empty list")
    resolved_command = [sys.executable if part == "{python}" else str(part) for part in command]
    cwd = _resolve(base_dir, str(when["cwd"])) if when.get("cwd") else project_dir
    timeout = int(when.get("timeout_seconds", 300))
    cmd_env, tmp_dir = _build_command_check_env(given, base_dir, project_dir, when=when)
    try:
        proc = subprocess.run(
            resolved_command,
            cwd=str(cwd),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
            env=cmd_env,
            check=False,
        )
    finally:
        if tmp_dir is not None:
            import shutil
            shutil.rmtree(tmp_dir, ignore_errors=True)
    expected_returncode = int(then.get("returncode", 0))
    stdout_contains = then.get("stdout_contains")
    stderr_contains = then.get("stderr_contains")
    ok = proc.returncode == expected_returncode
    if stdout_contains is not None:
        if isinstance(stdout_contains, list):
            ok = ok and all(s in proc.stdout for s in stdout_contains)
        else:
            ok = ok and str(stdout_contains) in proc.stdout
    if stderr_contains is not None:
        if isinstance(stderr_contains, list):
            ok = ok and all(s in proc.stderr for s in stderr_contains)
        else:
            ok = ok and str(stderr_contains) in proc.stderr
    return ok, "command matched expectations" if ok else "command did not match expectations", {
        "command": resolved_command,
        "cwd": str(cwd),
        "returncode": proc.returncode,
        "stdout": proc.stdout[-4000:],
        "stderr": proc.stderr[-4000:],
    }


SCORERS: dict[str, Callable[[dict[str, Any], Path, Path], tuple[bool, str, dict[str, Any]]]] = {
    "command_check": score_command_check,
    "csv_row_match": score_csv_row_match,
    "email_parse_check": score_email_parse_check,
    "excel_row_match": score_excel_row_match,
    "excel_schema_check": score_excel_schema_check,
    "exe_launch_check": score_exe_launch_check,
    "file_exists_check": score_file_exists_check,
    "file_output_check": score_file_output_check,
    "json_exact_match": score_json_exact_match,
    "mapping_rule_check": score_mapping_rule_check,
}
