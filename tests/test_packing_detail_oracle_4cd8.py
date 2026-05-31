"""Oracle tests for IMP-20260530-4CD8 — lookup_packing_dimensions sn 파라미터 검증.

이 테스트는 tests/oracles/IMP-20260530-4CD8/ 아래 oracle input/expected를 기준으로
실제 lookup_packing_dimensions(file_path, sn=...) 를 호출하여 반환값을 검증합니다.

테스트 구조:
  - 임시 xlsx 생성 (oracle input.json의 sheet_rows 스펙대로)
  - lookup_packing_dimensions(tmp_path, sn=...) 실제 호출
  - 반환값을 oracle expected.json과 비교
"""
import json
import os
import sys
import tempfile
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import pytest

# FEAT-20260426-332E를 import 경로에 추가
_REPO_ROOT = Path(__file__).parent.parent
_FEAT_DIR = _REPO_ROOT / "FEAT-20260426-332E"
if str(_FEAT_DIR) not in sys.path:
    sys.path.insert(0, str(_FEAT_DIR))

from core.packing_detail_reader import lookup_packing_dimensions  # noqa: E402

# --- Column constants (packing_detail_reader.py와 일치) ---
_COL_D = 4   # S/N
_COL_F = 6   # width
_COL_G = 7   # length/depth
_COL_H = 8   # height
_COL_J = 10  # weight
_COL_K = 11  # Crating Date

_ORACLE_DIR = _REPO_ROOT / "tests" / "oracles" / "IMP-20260530-4CD8"


def _build_temp_xlsx(sheet_rows: List[Dict[str, Any]]) -> str:
    """oracle input.json의 sheet_rows 스펙으로 임시 xlsx를 생성하고 경로를 반환합니다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2021년"  # type: ignore[assignment]

    # 헤더 행
    ws.cell(row=1, column=_COL_D, value="Project Id/S/N")
    ws.cell(row=1, column=_COL_F, value="F")
    ws.cell(row=1, column=_COL_G, value="G")
    ws.cell(row=1, column=_COL_H, value="H")
    ws.cell(row=1, column=_COL_J, value="J")
    ws.cell(row=1, column=_COL_K, value="Crating Date")

    for i, row_spec in enumerate(sheet_rows, start=2):
        ws.cell(row=i, column=_COL_D, value=row_spec.get("col_d"))
        ws.cell(row=i, column=_COL_F, value=row_spec.get("col_f"))
        ws.cell(row=i, column=_COL_G, value=row_spec.get("col_g"))
        ws.cell(row=i, column=_COL_H, value=row_spec.get("col_h"))
        ws.cell(row=i, column=_COL_J, value=row_spec.get("col_j"))
        k_val = row_spec.get("col_k_date")
        if k_val is not None:
            # ISO 날짜 문자열 → date 객체로 변환
            ws.cell(row=i, column=_COL_K, value=date.fromisoformat(k_val))
        else:
            ws.cell(row=i, column=_COL_K, value=None)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)
    wb.close()
    return tmp_path


def _load_oracle(case_dir_name: str):
    """oracle input.json과 expected.json을 로드합니다."""
    case_dir = _ORACLE_DIR / case_dir_name
    with open(case_dir / "input.json", encoding="utf-8") as f:
        inp = json.load(f)
    with open(case_dir / "expected.json", encoding="utf-8") as f:
        exp = json.load(f)
    return inp, exp


class TestOracleNormalSnMatch:
    """TC-1: normal case — sn이 D열에 있고 K열 날짜 유효 → 치수 문자열 반환."""

    def test_lookup_returns_dimension_string(self):
        inp, exp = _load_oracle("case_normal_sn_match")
        sn = inp["sn"]
        tmp_path = _build_temp_xlsx(inp["sheet_rows"])
        try:
            result = lookup_packing_dimensions(tmp_path, sn=sn)
            assert result is not None, f"None 반환 — sn='{sn}'가 D열에 있어야 합니다"
            assert result == exp["return_value"], (
                f"반환값 불일치: 실제='{result}' / 기대='{exp['return_value']}'"
            )
        finally:
            os.unlink(tmp_path)


class TestOracleEdgeSnNotFound:
    """TC-2: edge case — sn이 D열에 없으면 None 반환."""

    def test_lookup_returns_none_for_missing_sn(self):
        inp, exp = _load_oracle("case_edge_sn_not_found")
        sn = inp["sn"]
        tmp_path = _build_temp_xlsx(inp["sheet_rows"])
        try:
            result = lookup_packing_dimensions(tmp_path, sn=sn)
            assert result is None, (
                f"None이 아닌 값 반환 — sn='{sn}'는 D열에 없어야 합니다: '{result}'"
            )
            assert exp["return_value"] is None, "oracle expected.json이 null을 기대해야 합니다"
        finally:
            os.unlink(tmp_path)


class TestOracleEdgeNoKDate:
    """TC-3: edge case — sn이 D열에 있지만 K열 날짜 없음 → None 반환."""

    def test_lookup_returns_none_when_k_date_missing(self):
        inp, exp = _load_oracle("case_edge_no_k_date")
        sn = inp["sn"]
        tmp_path = _build_temp_xlsx(inp["sheet_rows"])
        try:
            result = lookup_packing_dimensions(tmp_path, sn=sn)
            assert result is None, (
                f"None이 아닌 값 반환 — K열 날짜 없는 행은 무시되어야 합니다: '{result}'"
            )
            assert exp["return_value"] is None, "oracle expected.json이 null을 기대해야 합니다"
        finally:
            os.unlink(tmp_path)


class TestCallerSnPassThrough:
    """TC-4/5: main.py와 ui/app.py 호출부가 sn을 실제로 전달하는지 코드 검증."""

    def test_main_py_passes_sn_not_project_id(self):
        """main.py의 lookup_packing_dimensions 호출이 project_id 인자를 사용하지 않는지 확인."""
        main_py = _FEAT_DIR / "main.py"
        content = main_py.read_text(encoding="utf-8")
        # project_id= 인자가 lookup_packing_dimensions 호출에 없어야 함
        assert "lookup_packing_dimensions" in content, "main.py에 lookup_packing_dimensions 호출 없음"
        # sn= 키워드 인자 또는 위치 인자로 sn 변수를 전달해야 함
        assert "project_id=" not in content.split("lookup_packing_dimensions")[1].split(")")[0], (
            "main.py가 lookup_packing_dimensions에 project_id= 인자를 전달하고 있습니다 — sn=으로 변경 필요"
        )

    def test_ui_app_py_passes_sn_not_project_id(self):
        """ui/app.py의 lookup_packing_dimensions 호출이 project_id 인자를 사용하지 않는지 확인."""
        app_py = _FEAT_DIR / "ui" / "app.py"
        content = app_py.read_text(encoding="utf-8")
        assert "lookup_packing_dimensions" in content, "ui/app.py에 lookup_packing_dimensions 호출 없음"
        # 모든 호출부에서 project_id= 인자 금지
        import re
        calls = re.findall(r"lookup_packing_dimensions\([^)]+\)", content)
        for call in calls:
            assert "project_id=" not in call, (
                f"ui/app.py가 lookup_packing_dimensions에 project_id= 전달: {call}"
            )
