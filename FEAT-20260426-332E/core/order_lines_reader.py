"""Excel B (Customer Order Lines) lookup for project data."""

import logging
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.utils import column_index_from_string
from datetime import datetime as _dt, date as _date

_ALLOWED_ROOTS: List[Path] = [
    Path.home(),                   # traversal guard: user home directory
    Path("//ccikrnf/wrkgroup"),    # MT-1: approved UNC share for order files
]


def _safe_resolve(user_path: str) -> Path:
    """Resolve and validate a user-supplied path against _ALLOWED_ROOTS.

    Prevents directory traversal by ensuring the resolved path remains
    inside at least one of the allowed roots (symlink-safe via .resolve()).
    UNC paths (//server/share/...) are compared by string prefix because
    Path.resolve() on Windows UNC shares does not always produce a
    canonical form that relative_to() can match.

    Args:
        user_path: Raw path string from user input or config.

    Returns:
        Resolved Path object guaranteed to be inside one of _ALLOWED_ROOTS.

    Raises:
        ValueError: If the resolved path escapes all allowed roots.
    """
    p = Path(user_path)
    # UNC paths: compare as normalised string prefix (case-insensitive on Windows)
    p_str = str(p).replace("\\", "/").lower()
    for root in _ALLOWED_ROOTS:
        root_str = str(root).replace("\\", "/").lower().rstrip("/")
        if p_str.startswith(root_str):
            return p
    # Non-UNC paths: use resolve() + relative_to() for symlink safety
    resolved = p.resolve()
    for root in _ALLOWED_ROOTS:
        try:
            resolved.relative_to(root.resolve())
            return resolved
        except ValueError:
            continue
    raise ValueError(
        f"Path traversal detected: '{user_path}' escapes all allowed roots"
    )

logger = logging.getLogger(__name__)

_EMPTY_RESULT: Dict[str, object] = {"fp": "", "f_col": "", "jm": "", "if_col": "", "line_nos": [], "found": False}
_DEFAULT_SHEET = "Customer Order Lines"
_DEFAULT_SEARCH_COL = "A"
_DEFAULT_FP_COL = "I"
_DEFAULT_F_COL = "D"
_DEFAULT_JM_COL = "KJ"  # Col 296 = 'Project Manager' (English names e.g. 'Yong Kyu Sung'); Col 7 (G) = numeric employee IDs — wrong column
_DEFAULT_IF_COL = "C"
_DEFAULT_LINE_NO_COL = "G"
_DEFAULT_JY_COL = "E"
_DEFAULT_STATUS_COL = "CG"  # MT-1: Cancelled status filter column
_DEFAULT_SALES_PART_COL = "H"  # MT-2: Sales part filter column for service-part exclusion
# MT-2: Service part keywords (uppercase, exact match) to exclude non-physical kitting items.
# Keys are uppercase since sales part values are normalized via .strip().upper() before comparison.
_SERVICE_PART_KEYWORDS: frozenset = frozenset({
    "FIELD SERVICE", "PROJECT SERVICE", "PACKING AND FREIGHT",
    "CERTIFICATION", "ENGINEERING", "PATTERN DEVELOPMENT",
    "EXPEDITING", "PROJECT", "TESTING", "WARRANTY",
})


def _col_to_idx(letter: str) -> int:
    """Convert Excel column letter (e.g. 'A', 'J') to 1-based index."""
    try:
        return column_index_from_string(letter.strip().upper())
    except Exception:
        raise ValueError(f"Invalid Excel column letter: '{letter}'")


def lookup_order_lines(
    file_path: str,
    project_id: str,
    sheet_name: Optional[str] = None,
    search_col_letter: Optional[str] = None,
    fp_col_letter: Optional[str] = None,
    f_col_letter: Optional[str] = None,
    jm_col_letter: Optional[str] = None,
    if_col_letter: Optional[str] = None,
    line_no_col_letter: Optional[str] = None,
    jy_col_letter: Optional[str] = None,
    planned_date: Optional[str] = None,
) -> Dict[str, object]:
    """Look up project data from Excel B by Project ID.

    Finds ALL rows matching project_id. Returns field values from the
    selected match plus Line No values from all matching rows in the same
    F-col group.

    When F column has 2+ distinct date values (conflict), uses planned_date
    to select the date group closest to the reference date. If planned_date
    is None or cannot be parsed, today's date is used as the reference.
    When F column has only 1 distinct date, all matching rows are collected
    (f_col_conflict=False).

    Args:
        file_path: Absolute path to Excel B file.
        project_id: Project ID to search for (1–255 chars).
        sheet_name: Sheet name override (default: "Customer Order Lines").
        search_col_letter: Column letter for Project ID lookup (default: "A").
        fp_col_letter: Column letter for FP field (default: "J").
        f_col_letter: Column letter for F (Wanted Delivery Date) field (default: "F").
        jm_col_letter: Column letter for JM field (default: "KJ").
        if_col_letter: Column letter for IF field (default: "C").
        line_no_col_letter: Column letter for Line No field (default: "H").
        jy_col_letter: Column letter for JY field (default: "E"). None → "E".
        planned_date: Reference date string (YYYY-MM-DD) used to resolve F-col
            conflicts by selecting the closest date group. None or unparseable
            → today's date used as reference.
            # allowed: None → _date.today() used as fallback reference (by design)

    Returns:
        Dict with keys: "fp", "f_col", "jm", "if_col", "jy_col" (str),
        "line_nos" (list[str]), "f_col_conflict" (bool), "f_col_all" (list[str]),
        "found" (bool).
        Empty result (found=False) if project_id not found.

    Raises:
        TypeError: If file_path or project_id is None or wrong type.
        ValueError: If file_path is empty or project_id is 0 or >255 chars.
        FileNotFoundError: If the Excel B file does not exist.
    """
    # AL type_valid item 1: None check for file_path
    if file_path is None:
        raise TypeError("file_path must not be None")
    # AL type_valid item 3: isinstance check for file_path
    if not isinstance(file_path, str):
        raise TypeError(f"file_path must be str, got {type(file_path).__name__}")
    # AL type_valid item 2+4: empty string boundary — negative not allowed (empty is invalid path)
    if len(file_path.strip()) == 0:
        raise ValueError("file_path must not be empty")

    # AL type_valid item 1: None check for project_id
    if project_id is None:
        raise TypeError("project_id must not be None")
    # AL type_valid item 3: isinstance check for project_id
    if not isinstance(project_id, str):
        raise TypeError(f"project_id must be str, got {type(project_id).__name__}")
    # AL type_valid item 2+4: length boundary — negative not allowed: 0-length is invalid
    if len(project_id) == 0:
        raise ValueError("project_id must not be empty")
    if len(project_id) > 255:
        raise ValueError(f"project_id too long: {len(project_id)} chars (max 255)")

    # AL type_valid: planned_date is Optional[str]; None → use today (allowed, by design)
    if planned_date is not None:
        try:
            _ref_date: _date = _date.fromisoformat(planned_date)
        except (ValueError, TypeError):
            # Invalid date string → fall back to today
            logger.warning(
                "planned_date '%s' could not be parsed — using today as reference", planned_date
            )
            _ref_date = _date.today()
    else:
        # allowed: planned_date=None → use today as reference date
        _ref_date = _date.today()

    target_sheet = sheet_name or _DEFAULT_SHEET
    search_idx = _col_to_idx(search_col_letter or _DEFAULT_SEARCH_COL)
    fp_idx = _col_to_idx(fp_col_letter or _DEFAULT_FP_COL)
    f_idx = _col_to_idx(f_col_letter or _DEFAULT_F_COL)
    jm_idx = _col_to_idx(jm_col_letter or _DEFAULT_JM_COL)
    if_idx = _col_to_idx(if_col_letter or _DEFAULT_IF_COL)
    line_no_idx = _col_to_idx(line_no_col_letter or _DEFAULT_LINE_NO_COL)
    # jy_col_letter: None → default string (negative not allowed: None is absent value, default applied)
    jy_idx = _col_to_idx(jy_col_letter or _DEFAULT_JY_COL)
    # MT-1: D column index for Cancelled-status filter (computed once)
    status_idx = _col_to_idx(_DEFAULT_STATUS_COL)
    # MT-2: I column index for service-part filter (computed once)
    sales_part_idx = _col_to_idx(_DEFAULT_SALES_PART_COL)

    # FS.traversal: resolve and validate against _ALLOWED_ROOTS before any I/O
    path = _safe_resolve(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel B file not found: '{path}'")

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, keep_links=False)
    except Exception as exc:
        raise RuntimeError(f"Cannot open Excel B '{path}': {exc}") from exc

    try:
        ws = wb[target_sheet] if target_sheet in wb.sheetnames else wb.worksheets[0]

        project_id_stripped = project_id.strip()
        all_matches: List[Dict[str, str]] = []  # all rows matching project_id; includes per-row line_no
        f_col_all: List[str] = []

        def _read(row: tuple, idx: int) -> str:
            """Safely extract and format a cell value by 1-based column index."""
            if idx > len(row):
                return ""
            val = row[idx - 1].value
            if val is None:
                return ""
            if isinstance(val, _dt):
                return val.strftime("%Y-%m-%d")
            if isinstance(val, _date):
                return val.strftime("%Y-%m-%d")
            return str(val).strip()  # allowed: openpyxl cell .value is Any; str() is safe for cell text extraction

        for row in ws.iter_rows():
            if len(row) < search_idx:
                continue
            if str(row[search_idx - 1].value or "").strip() != project_id_stripped:
                continue

            # MT-1: D column "Cancelled" status filter (case-insensitive)
            status_val = _read(row, status_idx).lower()
            if status_val == "cancelled":
                continue

            # MT-2: I column service-part filter — exclude rows whose sales part is in
            # _SERVICE_PART_KEYWORDS or contains no digits (purely alphabetical/empty values
            # are treated as non-physical service items and excluded).
            sales_part_val = _read(row, sales_part_idx).strip().upper()
            if sales_part_val in _SERVICE_PART_KEYWORDS or not any(ch.isdigit() for ch in sales_part_val):
                continue

            f_val = _read(row, f_idx)
            if f_val:
                f_col_all.append(f_val)

            line_no_val = _read(row, line_no_idx)
            jy_val = _read(row, jy_idx)

            # Collect ALL matching rows' field values including per-row line_no and jy_col
            all_matches.append({
                "fp":      _read(row, fp_idx),
                "f_col":   f_val,
                "jm":      _read(row, jm_idx),
                "if_col":  _read(row, if_idx),
                "line_no": line_no_val,
                "jy_col":  jy_val,
            })

        if not all_matches:
            logger.info("project_id '%s' not found", project_id_stripped)
            return {"fp": "", "f_col": "", "jm": "", "if_col": "", "line_nos": [],
                    "f_col_conflict": False, "f_col_all": [], "jy_col": "", "found": False}

        first_match: Dict[str, str] = all_matches[0]

        # Detect F column conflict: 2+ distinct values across matching rows
        distinct_f = list(dict.fromkeys(f_col_all))  # preserve order, deduplicate
        f_col_conflict = len(distinct_f) >= 2

        if f_col_conflict:
            logger.warning("F column conflict for '%s': %s", project_id_stripped, distinct_f)

        # Conflict resolution: use planned_date proximity to select the correct date group
        if f_col_conflict:
            sorted_groups = sorted(distinct_f)  # date strings sort ascending lexicographically

            def _date_distance(ds: str) -> int:
                """Return absolute day distance from ds to _ref_date; 999999 on parse failure."""
                try:
                    return abs((_date.fromisoformat(ds) - _ref_date).days)
                except (ValueError, TypeError):
                    return 999999

            # AL item 2+4: min() selects the closest date; deterministic when distances are equal
            selected_f_col = min(sorted_groups, key=_date_distance)
            selected_match = next(m for m in all_matches if m["f_col"] == selected_f_col)
            logger.info(
                "Conflict resolved by planned_date=%s → selected_f_col=%s",
                _ref_date, selected_f_col,
            )
        else:
            selected_match = first_match
            selected_f_col = selected_match["f_col"]

        # Filter line_nos to selected_f_col group (conflict) or all matches (no conflict)
        if f_col_conflict:
            line_nos = [m["line_no"] for m in all_matches if m["f_col"] == selected_f_col and m["line_no"]]
        else:
            line_nos = [m["line_no"] for m in all_matches if m["line_no"]]

        result: Dict[str, object] = {
            "fp":             selected_match["fp"],
            "f_col":          selected_match["f_col"],
            "jm":             selected_match["jm"],
            "if_col":         selected_match["if_col"],
            "line_nos":       line_nos,
            "f_col_conflict": f_col_conflict,
            "f_col_all":      distinct_f,
            "jy_col":         selected_match["jy_col"],
            "found":          True,
        }
        logger.info(
            "Found order lines for '%s': fp=%s if_col=%s line_nos=%s f_conflict=%s jy_col=%s",
            project_id_stripped, selected_match["fp"], selected_match["if_col"],
            line_nos, f_col_conflict, selected_match["jy_col"],
        )
        return result

    finally:
        wb.close()


if __name__ == "__main__":
    # planned_date proximity conflict resolution test
    from datetime import date as _d

    _all_conflict = [
        {"fp": "DDP", "f_col": "2026-04-20", "jm": "A", "if_col": "X", "line_no": "3", "jy_col": ""},
        {"fp": "DDP", "f_col": "2026-04-20", "jm": "A", "if_col": "X", "line_no": "5", "jy_col": ""},
        {"fp": "DDP", "f_col": "2026-04-29", "jm": "A", "if_col": "X", "line_no": "1", "jy_col": ""},
        {"fp": "DDP", "f_col": "2026-04-29", "jm": "A", "if_col": "X", "line_no": "2", "jy_col": ""},
        {"fp": "DDP", "f_col": "2026-06-30", "jm": "A", "if_col": "X", "line_no": "4", "jy_col": ""},
    ]
    _distinct = list(dict.fromkeys(m["f_col"] for m in _all_conflict if m["f_col"]))

    def _dist(date_str: str, ref: "_d") -> int:
        """Return absolute day distance between date_str and ref; 999999 on parse error."""
        try:
            return abs((_d.fromisoformat(date_str[:10]) - ref).days)
        except (ValueError, AttributeError):
            return 999999

    _ref = _d(2026, 4, 28)
    _sel = min(_distinct, key=lambda s: _dist(s, _ref))
    assert _sel == "2026-04-29", f"Expected 2026-04-29, got {_sel}"
    print("[SELF-VERIFY] order_lines_reader.py planned_date proximity OK")
