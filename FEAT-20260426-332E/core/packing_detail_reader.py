"""Packing detail lookup from the '패킹 디테일' Excel file.

Opens the Excel file in read_only + data_only mode (bypasses file locks —
openpyxl reads ZIP bytes directly; data_only=True returns cached formula
values instead of formula strings).

Search key : Column D (S/N)
Date filter : Column K (Crating Date) — rows with no valid K date are excluded;
              when multiple rows match, the one with K date closest to today is selected.
Dimensions  : Columns F, G, H, J
"""

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

logger = logging.getLogger(__name__)

# Column indices (1-based)
_COL_PROJ_ID = 4      # Column D — "Project Id/S/N"
_COL_F = 6            # Column F — width
_COL_G = 7            # Column G — length/depth
_COL_H = 8            # Column H — height
_COL_J = 10           # Column J — weight
_COL_CRATING_DATE = 11  # Column K — "Crating Date"

_DEFAULT_SHEET = "2021년"


def lookup_packing_dimensions(
    file_path: str,
    sn: str,
    sheet_name: Optional[str] = None,
) -> Optional[str]:
    """Look up packing dimensions for a given S/N from the packing detail Excel.

    Opens the file with openpyxl read_only=True, data_only=True (no COM, no GUI
    popup; data_only returns cached VLOOKUP values). Searches column D (S/N)
    for sn. Rows with no valid K (Crating Date) are excluded.
    When multiple rows match, the row whose K date is closest to date.today()
    is selected automatically (no conflict list is returned).

    Return semantics (2 states):
        None  — 0 rows with a valid K date: note should be "포장반" (no dimensions).
        str   — 1 or more rows; the one closest to today is returned as a formatted
                dimension string e.g. "55*55*20 8.0kg".

    Args:
        file_path: Absolute or relative path to the packing detail Excel file.
        sn: The S/N value to search for in column D.
        sheet_name: Sheet to search in. Defaults to "2021년".

    Returns:
        None or a dimension string.

    Raises:
        TypeError: If file_path or sn is None or wrong type.
        ValueError: If file_path or sn is an empty string.
        FileNotFoundError: If the Excel file does not exist at file_path.
        RuntimeError: If the workbook cannot be opened.
    """
    # --- AL: None → isinstance checks (fixed order) ---
    if file_path is None:
        raise TypeError("file_path must not be None")
    if not isinstance(file_path, str):
        raise TypeError(f"file_path must be str, got {type(file_path).__name__}")
    if len(file_path.strip()) == 0:
        raise ValueError("file_path must not be empty")

    if sn is None:
        raise TypeError("sn must not be None")
    if not isinstance(sn, str):
        raise TypeError(f"sn must be str, got {type(sn).__name__}")
    if len(sn.strip()) == 0:
        raise ValueError("sn must not be empty")
    # sheet_name: None allowed (triggers default); no negative numeric value possible for str — no boundary check needed.

    path = Path(file_path)
    if not path.exists():
        logger.error("Packing detail file not found: '%s'", path)
        raise FileNotFoundError(f"Packing detail file not found: '{path}'")

    logger.info("Opening packing detail file (read_only, data_only): '%s'", path)

    try:
        wb = openpyxl.load_workbook(
            str(path), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:
        logger.error("Failed to open packing detail file '%s': %s", path, exc)
        raise RuntimeError(
            f"Cannot open packing detail file '{path}': {exc}"
        ) from exc

    try:
        # Resolve worksheet
        target_sheet = sheet_name if sheet_name else _DEFAULT_SHEET
        if target_sheet in wb.sheetnames:
            ws = wb[target_sheet]
            logger.info("Using sheet '%s'", target_sheet)
        else:
            ws = wb.worksheets[0]
            logger.warning(
                "Sheet '%s' not found; using first sheet '%s'",
                target_sheet,
                ws.title,
            )

        sn_stripped = sn.strip()
        matches: List[Dict[str, object]] = []

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            # Column D check (1-based index _COL_PROJ_ID = 4 → 0-based = 3)
            if len(row) < _COL_PROJ_ID:
                continue
            cell_d = row[_COL_PROJ_ID - 1]
            d_val = cell_d.value
            if d_val is None:
                continue
            parts = [p.strip() for p in str(d_val).strip().split('/')]
            if sn_stripped not in parts:
                continue

            # Column K date filter (1-based _COL_CRATING_DATE = 11 → 0-based = 10)
            if len(row) < _COL_CRATING_DATE:
                # K column absent in this row — skip (not yet scheduled)
                continue
            k_val = row[_COL_CRATING_DATE - 1].value
            if k_val is None:
                # No date set — skip (not yet scheduled / done)
                continue

            # Normalise to date object
            if isinstance(k_val, datetime):
                k_date: date = k_val.date()
            elif isinstance(k_val, date):
                k_date = k_val
            else:
                # Try parsing as ISO string (e.g. "2026-04-30" entered as text)
                try:
                    k_date = date.fromisoformat(str(k_val).strip())
                except (ValueError, AttributeError):
                    # Fallback: MM/DD/YY format (e.g. "03/18/26" entered as text)
                    try:
                        k_date = datetime.strptime(str(k_val).strip(), "%m/%d/%y").date()
                    except (ValueError, AttributeError):
                        continue  # Not a parseable date — skip

            # No target_date comparison; all rows with a valid K date are candidates.

            # Build dimension string from columns F, G, H, J
            def _cell_str(col_1based: int) -> str:
                """Return cell value as stripped string, or '' if absent/None."""
                if len(row) < col_1based:
                    return ""
                val = row[col_1based - 1].value
                return str(val).strip() if val is not None else ""

            f_val = _cell_str(_COL_F)
            g_val = _cell_str(_COL_G)
            h_val = _cell_str(_COL_H)
            j_val = _cell_str(_COL_J)
            dims = f"{f_val}*{g_val}*{h_val} {j_val}kg"

            matches.append({"row": row_idx, "dims": dims, "date": k_date})

        match_count = len(matches)
        logger.info(
            "sn='%s' → %d matching row(s) with valid K date",
            sn_stripped, match_count,
        )

        if match_count == 0:
            return None
        if match_count == 1:
            return str(matches[0]["dims"])
        # 2+ matches — select the row with K date closest to today
        today = date.today()
        best = min(
            matches,
            key=lambda m: abs(
                (m["date"].date() if isinstance(m["date"], datetime) else m["date"]) - today  # allowed: m["date"] is date|datetime, stored from validated isinstance block above
            ),
        )
        logger.info(
            "sn='%s' — %d rows found, selected closest to today(%s): row=%s date=%s",
            sn_stripped, match_count, today, best["row"], best["date"],
        )
        return str(best["dims"])

    finally:
        wb.close()


if __name__ == "__main__":
    import os
    import tempfile
    import openpyxl as xl
    from datetime import date as _date

    _today = _date.today()

    # --- Build minimal test workbook ---
    wb_test = xl.Workbook()
    ws_test = wb_test.active
    ws_test.title = "2021년"

    # Header row (row 1) — not data
    ws_test.cell(row=1, column=_COL_PROJ_ID, value="Project Id/S/N")
    ws_test.cell(row=1, column=_COL_F, value="F")
    ws_test.cell(row=1, column=_COL_G, value="G")
    ws_test.cell(row=1, column=_COL_H, value="H")
    ws_test.cell(row=1, column=_COL_J, value="J")
    ws_test.cell(row=1, column=_COL_CRATING_DATE, value="Crating Date")

    # Row 2: PROJ-001 — single row with a valid K date
    ws_test.cell(row=2, column=_COL_PROJ_ID, value="PROJ-001")
    ws_test.cell(row=2, column=_COL_F, value="55")
    ws_test.cell(row=2, column=_COL_G, value="55")
    ws_test.cell(row=2, column=_COL_H, value="20")
    ws_test.cell(row=2, column=_COL_J, value="8.0")
    ws_test.cell(row=2, column=_COL_CRATING_DATE, value=_today)

    # Row 3: PROJ-002 — K=None → should be excluded → result None
    ws_test.cell(row=3, column=_COL_PROJ_ID, value="PROJ-002")
    ws_test.cell(row=3, column=_COL_F, value="30")
    ws_test.cell(row=3, column=_COL_G, value="30")
    ws_test.cell(row=3, column=_COL_H, value="30")
    ws_test.cell(row=3, column=_COL_J, value="5.0")
    # No K value set → None

    # Rows 4 & 5: PROJ-MULTI — two rows with valid K dates; today+1 is closer to today than today-30
    _near_date = _date(_today.year, _today.month, _today.day)   # same day = distance 0
    _far_date = _date(2020, 1, 1)                                 # very old date
    ws_test.cell(row=4, column=_COL_PROJ_ID, value="PROJ-MULTI")
    ws_test.cell(row=4, column=_COL_F, value="10")
    ws_test.cell(row=4, column=_COL_G, value="10")
    ws_test.cell(row=4, column=_COL_H, value="10")
    ws_test.cell(row=4, column=_COL_J, value="1.0")
    ws_test.cell(row=4, column=_COL_CRATING_DATE, value=_near_date)

    ws_test.cell(row=5, column=_COL_PROJ_ID, value="PROJ-MULTI")
    ws_test.cell(row=5, column=_COL_F, value="99")
    ws_test.cell(row=5, column=_COL_G, value="99")
    ws_test.cell(row=5, column=_COL_H, value="99")
    ws_test.cell(row=5, column=_COL_J, value="9.0")
    ws_test.cell(row=5, column=_COL_CRATING_DATE, value=_far_date)

    # Row 6: slash-combined D col 'SNPART_1st/PROJ-SLASH'
    ws_test.cell(row=6, column=_COL_PROJ_ID, value="SNPART_1st/PROJ-SLASH")
    ws_test.cell(row=6, column=_COL_F, value="20")
    ws_test.cell(row=6, column=_COL_G, value="30")
    ws_test.cell(row=6, column=_COL_H, value="40")
    ws_test.cell(row=6, column=_COL_J, value="5.5")
    ws_test.cell(row=6, column=_COL_CRATING_DATE, value=_today)

    # Row 7: MM/DD/YY K date string
    ws_test.cell(row=7, column=_COL_PROJ_ID, value="PROJ-MMDDYY")
    ws_test.cell(row=7, column=_COL_F, value="11")
    ws_test.cell(row=7, column=_COL_G, value="22")
    ws_test.cell(row=7, column=_COL_H, value="33")
    ws_test.cell(row=7, column=_COL_J, value="4.4")
    ws_test.cell(row=7, column=_COL_CRATING_DATE, value="04/29/26")  # MM/DD/YY string

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    wb_test.save(tmp_path)
    wb_test.close()

    try:
        # --- Test 1: exactly 1 match → str ---
        result = lookup_packing_dimensions(tmp_path, "PROJ-001")
        assert isinstance(result, str), f"Expected str, got {type(result)}: {result}"
        assert result == "55*55*20 8.0kg", f"Wrong dims: '{result}'"

        # --- Test 2: K=None row → None ---
        result2 = lookup_packing_dimensions(tmp_path, "PROJ-002")
        assert result2 is None, f"Expected None (K=None), got {result2}"

        # --- Test 3: no match at all ---
        result3 = lookup_packing_dimensions(tmp_path, "PROJ-NOTEXIST")
        assert result3 is None, f"Expected None for missing sn, got {result3}"

        # --- Test 4: multiple rows → closest-to-today selected (near_date wins over far_date) ---
        result4 = lookup_packing_dimensions(tmp_path, "PROJ-MULTI")
        assert isinstance(result4, str), f"Expected str (closest row), got {type(result4)}"
        assert result4 == "10*10*10 1.0kg", f"Expected near row dims, got '{result4}'"

        # --- Test 5: None inputs → TypeError ---
        try:
            lookup_packing_dimensions(None, "PROJ")  # type: ignore[arg-type]
            assert False, "Expected TypeError for None file_path"
        except TypeError:
            pass

        try:
            lookup_packing_dimensions(tmp_path, None)  # type: ignore[arg-type]
            assert False, "Expected TypeError for None sn"
        except TypeError:
            pass

        # --- Test 6: empty strings → ValueError ---
        try:
            lookup_packing_dimensions("   ", "PROJ")
            assert False, "Expected ValueError for empty file_path"
        except ValueError:
            pass

        try:
            lookup_packing_dimensions(tmp_path, "   ")
            assert False, "Expected ValueError for empty sn"
        except ValueError:
            pass

        # --- Test 7: wrong type for file_path → TypeError ---
        try:
            lookup_packing_dimensions(12345, "PROJ")  # type: ignore[arg-type]
            assert False, "Expected TypeError for int file_path"
        except TypeError:
            pass

        # --- Test 8: missing file → FileNotFoundError ---
        try:
            lookup_packing_dimensions("/nonexistent/file.xlsx", "PROJ")
            assert False, "Expected FileNotFoundError"
        except FileNotFoundError:
            pass

        # --- Test 9: D column slash-combined format matching ---
        result9 = lookup_packing_dimensions(tmp_path, "PROJ-SLASH")
        assert isinstance(result9, str), f"Expected str for slash-combined D, got {type(result9)}: {result9}"
        assert result9 == "20*30*40 5.5kg", f"Wrong dims for slash-combined D: '{result9}'"

        # --- Test 10: K column MM/DD/YY string date parsing ---
        result10 = lookup_packing_dimensions(tmp_path, "PROJ-MMDDYY")
        assert isinstance(result10, str), f"Expected str for MM/DD/YY K date, got {type(result10)}: {result10}"
        assert result10 == "11*22*33 4.4kg", f"Wrong dims for MM/DD/YY K date: '{result10}'"

    finally:
        os.unlink(tmp_path)

    print("[SELF-VERIFY] packing_detail_reader.py OK")
