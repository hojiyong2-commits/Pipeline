"""Excel A writer for AFM-Kitting mapped rows."""

import logging
import os
import shutil
import tempfile
import time as _time
from pathlib import Path
from typing import Dict, Optional

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string

logger = logging.getLogger(__name__)

_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# New column layout: A=Month, B=Date, C=구분, D=PM, E=SN, F=Location,
#                    G=PO, H=Project ID, I=Contract, J=Incoterm, K=Note
_DEFAULT_COLUMN_MAP: dict = {
    "month":      "A",
    "day":        "B",
    "division":   "C",
    "pm":         "D",
    "sn":         "E",
    "location":   "F",
    "po":         "G",
    "project_id": "H",
    "contract":   "I",
    "incoterm":   "J",
    "note":       "K",
}


def _apply_label_via_excel(file_path: str) -> None:
    """Open file with a hidden Excel instance to apply sensitivity label, then save and close."""
    try:
        import win32com.client as _win32
        xl = _win32.DispatchEx("Excel.Application")  # new instance, never connects to open Excel
        xl.Visible = False
        xl.DisplayAlerts = False
        try:
            wb = xl.Workbooks.Open(file_path)
            wb.Save()
            wb.Close(False)
        finally:
            xl.Quit()
    except Exception as exc:
        logger.warning("레이블 적용 실패 (무시하고 계속): %s", exc)


def _find_first_empty_row(ws, anchor_col_idx: int) -> int:
    """Find the first row where the anchor column cell is empty, starting from row 2."""
    row_idx = 2
    max_row = ws.max_row + 1 if ws.max_row else 2
    while row_idx <= max_row + 1:
        val = ws.cell(row=row_idx, column=anchor_col_idx).value
        if val is None or str(val).strip() == "":
            return row_idx
        row_idx += 1
    return row_idx


def write_to_excel_a(
    output_path: str,
    rows: list,
    column_map: Optional[dict] = None,
) -> None:
    """Write MappedRow entries into the existing Excel A file atomically.

    New column layout:
        A=Month, B=Date, C=구분("IC"), D=PM, E=SN, F=Location,
        G=PO, H=Project ID, I=Contract, J=Incoterm, K=Note

    Cell styles of existing rows are never modified.
    Values are written only; no format changes.

    Raises:
        TypeError: If output_path is None or not str; if rows is None.
        ValueError: If output_path is empty.
        FileNotFoundError: If Excel A does not exist.
        RuntimeError: If write fails.
    """
    if output_path is None:
        raise TypeError("output_path must not be None")
    if not isinstance(output_path, str):
        raise TypeError(f"output_path must be str, got {type(output_path).__name__}")
    if len(output_path.strip()) == 0:
        raise ValueError("output_path must not be empty")
    if rows is None:
        raise TypeError("rows must not be None")
    if not isinstance(rows, list):
        raise TypeError(f"rows must be list, got {type(rows).__name__}")

    col_map = _DEFAULT_COLUMN_MAP.copy()
    if column_map is not None:
        if not isinstance(column_map, dict):
            raise TypeError(f"column_map must be dict or None, got {type(column_map).__name__}")
        col_map.update(column_map)

    path = Path(output_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel A file not found: '{path}'")

    if not rows:
        logger.info("No rows to write — skipping Excel A update")
        return

    try:
        wb = openpyxl.load_workbook(str(path), keep_links=False)
    except Exception as exc:
        raise RuntimeError(f"Cannot open Excel A '{path}': {exc}") from exc

    try:
        ws = wb.active

        month_col = column_index_from_string(col_map.get("month",      "A"))
        day_col   = column_index_from_string(col_map.get("day",        "B"))
        div_col   = column_index_from_string(col_map.get("division",   "C"))
        pm_col    = column_index_from_string(col_map.get("pm",         "D"))
        sn_col    = column_index_from_string(col_map.get("sn",         "E"))
        loc_col   = column_index_from_string(col_map.get("location",   "F"))
        po_col    = column_index_from_string(col_map.get("po",         "G"))
        pid_col   = column_index_from_string(col_map.get("project_id", "H"))
        con_col   = column_index_from_string(col_map.get("contract",   "I"))
        inc_col   = column_index_from_string(col_map.get("incoterm",   "J"))
        note_col  = column_index_from_string(col_map.get("note",       "K"))

        # Build map of existing SNs → row numbers (col E used as duplicate key)
        existing_sn_rows: Dict[str, int] = {}
        max_used_row = 1
        for r_idx in range(2, (ws.max_row or 1) + 1):
            sn_val = ws.cell(row=r_idx, column=sn_col).value
            if sn_val is not None and str(sn_val).strip():
                existing_sn_rows[str(sn_val).strip()] = r_idx
                max_used_row = max(max_used_row, r_idx)

        # Find first truly empty row for appending (skip gaps)
        next_append_row = _find_first_empty_row(ws, month_col)

        logger.info("Writing %d rows (skip-if-exists); existing SNs in sheet: %d",
                    len(rows), len(existing_sn_rows))

        appended = 0
        for row in rows:
            sn = str(row.get("sn", "")).strip()
            if sn and sn in existing_sn_rows:
                logger.info("Skip row: sn='%s' already exists in Excel A (row %d) — not modified",
                            sn, existing_sn_rows[sn])
                continue
            r = next_append_row + appended
            appended += 1

            ws.cell(row=r, column=month_col).value = row.get("month", "")
            ws.cell(row=r, column=day_col).value   = row.get("day", "")
            ws.cell(row=r, column=div_col).value   = "IC"
            ws.cell(row=r, column=pm_col).value    = row.get("pm", "")
            ws.cell(row=r, column=sn_col).value    = row.get("sn", "")
            ws.cell(row=r, column=loc_col).value   = row.get("location", "")
            ws.cell(row=r, column=po_col).value    = row.get("po", "")
            ws.cell(row=r, column=pid_col).value   = row.get("project_id", "")
            ws.cell(row=r, column=con_col).value   = row.get("contract", "")
            ws.cell(row=r, column=inc_col).value   = row.get("incoterm", "")
            ws.cell(row=r, column=note_col).value  = row.get("note", "")
            for _col in (month_col, day_col, div_col, pm_col, sn_col,
                         loc_col, po_col, pid_col, con_col, inc_col, note_col):
                ws.cell(row=r, column=_col).alignment = _CENTER_ALIGN
            logger.info("Appended row %d: project_id=%s location=%s",
                        r, row.get("project_id", ""), row.get("location", ""))

        # After all appends: re-sort all data rows by PM
        _all_rows_data: list = []
        _sort_end_row = (next_append_row + appended - 1) if appended > 0 else max_used_row
        for _r in range(2, max(_sort_end_row, max_used_row) + 1):
            _pm_val = ws.cell(row=_r, column=pm_col).value
            if _pm_val is None or str(_pm_val).strip() == "":  # allowed: openpyxl cell .value is Any; str() safe for blank-check
                continue
            _all_rows_data.append({
                "month":      ws.cell(row=_r, column=month_col).value,
                "day":        ws.cell(row=_r, column=day_col).value,
                "division":   ws.cell(row=_r, column=div_col).value,
                "pm":         ws.cell(row=_r, column=pm_col).value,
                "sn":         ws.cell(row=_r, column=sn_col).value,
                "location":   ws.cell(row=_r, column=loc_col).value,
                "po":         ws.cell(row=_r, column=po_col).value,
                "project_id": ws.cell(row=_r, column=pid_col).value,
                "contract":   ws.cell(row=_r, column=con_col).value,
                "incoterm":   ws.cell(row=_r, column=inc_col).value,
                "note":       ws.cell(row=_r, column=note_col).value,
            })

        _all_rows_data.sort(key=lambda x: str(x.get("pm") or "").lower())  # allowed: pm value is Any; str() safe for sort key

        for _i, _rd in enumerate(_all_rows_data):
            _r = 2 + _i
            ws.cell(row=_r, column=month_col).value = _rd["month"]
            ws.cell(row=_r, column=day_col).value   = _rd["day"]
            ws.cell(row=_r, column=div_col).value   = _rd["division"]
            ws.cell(row=_r, column=pm_col).value    = _rd["pm"]
            ws.cell(row=_r, column=sn_col).value    = _rd["sn"]
            ws.cell(row=_r, column=loc_col).value   = _rd["location"]
            ws.cell(row=_r, column=po_col).value    = _rd["po"]
            ws.cell(row=_r, column=pid_col).value   = _rd["project_id"]
            ws.cell(row=_r, column=con_col).value   = _rd["contract"]
            ws.cell(row=_r, column=inc_col).value   = _rd["incoterm"]
            ws.cell(row=_r, column=note_col).value  = _rd["note"]
            for _col in (month_col, day_col, div_col, pm_col, sn_col,
                         loc_col, po_col, pid_col, con_col, inc_col, note_col):
                ws.cell(row=_r, column=_col).alignment = _CENTER_ALIGN

        # Clear trailing ghost rows (rows beyond sorted data range that may hold stale data)
        _clear_start = 2 + len(_all_rows_data)
        _clear_end = max(max_used_row, _sort_end_row)
        if _clear_start <= _clear_end:
            for _r in range(_clear_start, _clear_end + 1):
                for _col in (month_col, day_col, div_col, pm_col, sn_col,
                             loc_col, po_col, pid_col, con_col, inc_col, note_col):
                    ws.cell(row=_r, column=_col).value = None
            logger.info("Cleared %d trailing ghost rows (%d~%d)",
                        _clear_end - _clear_start + 1, _clear_start, _clear_end)

        logger.info("Excel A rows re-sorted by PM: total %d data rows", len(_all_rows_data))

        tmp_path: Optional[str] = None
        try:
            with tempfile.NamedTemporaryFile(
                    suffix=".xlsx", delete=False,
                    dir=tempfile.gettempdir()) as _f:
                tmp_path = _f.name
            wb.save(tmp_path)
            _apply_label_via_excel(tmp_path)
            _MAX_RETRY = 5
            for _attempt in range(1, _MAX_RETRY + 1):
                try:
                    shutil.copy2(tmp_path, str(path))
                    break
                except (PermissionError, OSError) as _copy_exc:
                    if _attempt == _MAX_RETRY:
                        raise
                    logger.warning(
                        "Excel A 저장 재시도 %d/%d: %s", _attempt, _MAX_RETRY, _copy_exc
                    )
                    _time.sleep(1)
        except Exception as exc:
            raise RuntimeError(f"Failed to write Excel A to '{path}': {exc}") from exc
        finally:
            if tmp_path is not None:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

    finally:
        wb.close()

    logger.info("Excel A updated: %d appended, %d skipped (already existed)", appended, len(rows) - appended)


def update_note_by_sn(
    output_path: str,
    sn: str,
    new_note: str,
    column_map: Optional[dict] = None,
) -> bool:
    """Find a row in Excel A by SN (col E) and update its Note (col K).

    Searches the active worksheet for a row whose SN column (default col E)
    matches the given sn, then overwrites the Note column (default col K)
    with new_note and saves directly to the file path.

    Args:
        output_path: Absolute path to the Excel A file (str).
        sn: Serial number to search for.
        new_note: New note value to write (empty string is allowed — intentional blank).
        column_map: Optional dict to override default column letters.

    Returns:
        True if the SN was found and the note updated; False if SN not found
        or file does not exist.

    Raises:
        TypeError: If output_path, sn, or new_note is None or wrong type;
                   if column_map is not dict or None.
        ValueError: If output_path or sn is empty/whitespace-only.
        RuntimeError: If saving the updated workbook fails (chained).
    """
    # AL item 1: None guard — output_path
    if output_path is None:
        raise TypeError("output_path must not be None")
    # AL item 3: isinstance type guard — output_path
    if not isinstance(output_path, str):
        raise TypeError(
            f"output_path must be str, got {type(output_path).__name__}"
        )
    # AL item 2 & 4: empty string is a boundary case — string length must be >= 1 non-whitespace
    if len(output_path.strip()) == 0:
        raise ValueError("output_path must not be empty")

    # AL item 1: None guard — sn
    if sn is None:
        raise TypeError("sn must not be None")
    # AL item 3: isinstance type guard — sn
    if not isinstance(sn, str):
        raise TypeError(f"sn must be str, got {type(sn).__name__}")
    # AL item 2 & 4: empty sn is not meaningful — length must be >= 1 non-whitespace
    if len(sn.strip()) == 0:
        raise ValueError("sn must not be empty")

    # AL item 1: None guard — new_note
    if new_note is None:
        raise TypeError("new_note must not be None")
    # AL item 3: isinstance type guard — new_note
    if not isinstance(new_note, str):
        raise TypeError(f"new_note must be str, got {type(new_note).__name__}")
    # AL item 4: empty new_note IS allowed — intentional blank note clears the cell

    col_map = _DEFAULT_COLUMN_MAP.copy()
    if column_map is not None:
        # AL item 3: isinstance type guard — column_map
        if not isinstance(column_map, dict):
            raise TypeError(
                f"column_map must be dict or None, got {type(column_map).__name__}"
            )
        col_map.update(column_map)

    path = Path(output_path)
    if not path.exists():
        logger.warning(
            "update_note_by_sn: 파일 미존재 '%s' — False 반환", path
        )
        return False

    try:
        wb = openpyxl.load_workbook(str(path), keep_links=False)
    except Exception as exc:
        logger.error(
            "update_note_by_sn: 파일 열기 실패 '%s': %s", path, exc
        )
        return False

    found = False
    try:
        ws = wb.active
        sn_col = column_index_from_string(col_map.get("sn", "E"))
        note_col = column_index_from_string(col_map.get("note", "K"))

        MAX_ROW = 10_000  # 병리적 입력 방어 상한 — while True 금지
        for r_idx in range(2, min((ws.max_row or 1) + 1, MAX_ROW + 1)):
            cell_val = ws.cell(row=r_idx, column=sn_col).value
            # allowed: openpyxl cell .value is Any; str() is safe for SN comparison
            if cell_val is not None and str(cell_val).strip() == sn.strip():
                ws.cell(row=r_idx, column=note_col).value = new_note
                ws.cell(row=r_idx, column=note_col).alignment = _CENTER_ALIGN
                found = True
                logger.info(
                    "Note 업데이트: sn='%s' row=%d note='%s'",
                    sn, r_idx, new_note,
                )
                break

        if not found:
            return False

        tmp_path: Optional[str] = None
        try:
            with tempfile.NamedTemporaryFile(
                    suffix=".xlsx", delete=False,
                    dir=tempfile.gettempdir()) as _f:
                tmp_path = _f.name
            wb.save(tmp_path)
            _apply_label_via_excel(tmp_path)
            _MAX_RETRY_NOTE = 5
            for _attempt_note in range(1, _MAX_RETRY_NOTE + 1):
                try:
                    shutil.copy2(tmp_path, str(path))
                    break
                except (PermissionError, OSError) as _copy_exc_note:
                    if _attempt_note == _MAX_RETRY_NOTE:
                        raise
                    logger.warning(
                        "update_note_by_sn 저장 재시도 %d/%d: %s",
                        _attempt_note, _MAX_RETRY_NOTE, _copy_exc_note
                    )
                    _time.sleep(1)
        except Exception as exc:
            raise RuntimeError(
                f"update_note_by_sn 저장 실패 '{path}': {exc}"
            ) from exc
        finally:
            if tmp_path is not None:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

    finally:
        wb.close()

    return found


def read_note_by_sn(
    output_path: str,
    sn: str,
    column_map: Optional[dict] = None,
) -> Optional[str]:
    """Read the current Note cell value for a given SN from Excel A.

    Returns the cell value as str, or None if the SN is not found or the file
    does not exist. Never raises on missing file or missing SN.

    Raises:
        TypeError: If output_path or sn is None or wrong type.
        ValueError: If output_path or sn is empty.
    """
    if output_path is None:
        raise TypeError("output_path must not be None")
    if not isinstance(output_path, str):
        raise TypeError(f"output_path must be str, got {type(output_path).__name__}")
    if len(output_path.strip()) == 0:
        raise ValueError("output_path must not be empty")
    if sn is None:
        raise TypeError("sn must not be None")
    if not isinstance(sn, str):
        raise TypeError(f"sn must be str, got {type(sn).__name__}")
    if len(sn.strip()) == 0:
        raise ValueError("sn must not be empty")

    col_map = _DEFAULT_COLUMN_MAP.copy()
    if column_map is not None:
        if not isinstance(column_map, dict):
            raise TypeError(
                f"column_map must be dict or None, got {type(column_map).__name__}"
            )
        col_map.update(column_map)

    path = Path(output_path)
    if not path.exists():
        return None

    try:
        wb = openpyxl.load_workbook(
            str(path), keep_links=False, read_only=True, data_only=True
        )
    except Exception:
        return None

    try:
        ws = wb.active
        sn_col = column_index_from_string(col_map.get("sn", "E"))
        note_col = column_index_from_string(col_map.get("note", "K"))
        sn_stripped = sn.strip()
        for row in ws.iter_rows(min_row=2):
            if len(row) < max(sn_col, note_col):
                continue
            cell_val = row[sn_col - 1].value
            if cell_val is not None and str(cell_val).strip() == sn_stripped:  # allowed: openpyxl cell .value is Any; str() for comparison
                note_val = row[note_col - 1].value
                return str(note_val) if note_val is not None else ""  # allowed: openpyxl cell .value is Any; str() safe for return
        return None
    finally:
        wb.close()
